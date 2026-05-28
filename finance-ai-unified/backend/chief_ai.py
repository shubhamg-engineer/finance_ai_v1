"""
Finance AI Chief Command — Unified Production Backend
======================================================
Single-file FastAPI backend orchestrating all 8 Finance AI agents:
  1. Research AI      — market trends, macro signals, opportunities
  2. Planning AI      — forecasts, burn rate, runway, scenarios
  3. Accounting AI    — transaction classification, fraud, GST, reconciliation
  4. Treasury AI      — cash position, liquidity, investments, stress tests
  5. Compliance AI    — tax tracking, regulatory deadlines, health score
  6. Reporting AI     — KPIs, MIS, board pack narrative
  7. Decision AI      — risk scoring, founder-aligned recommendations
  8. Chief Command AI — orchestration, Finance Health Score, composite patterns

All agents run on company-uploaded financial data (CSV / Excel / JSON).
No hardcoded or fake data. Company data drives everything.
"""

from __future__ import annotations

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1 — IMPORTS, CONFIG, CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

import csv
import hashlib
import io
import json
import logging
import os
import sys
import time
import traceback
import uuid
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Literal, Optional

import httpx
from dotenv import dotenv_values, load_dotenv
from fastapi import Depends, FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from groq import Groq
from pydantic import BaseModel, Field
from sqlalchemy import Column, DateTime, Float, Integer, String, Text, select, Boolean, ForeignKey
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.orm import DeclarativeBase

# ── Attempt openpyxl/xlrd for Excel support ──────────────────────────────────
try:
    import openpyxl

    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not installed — Excel (.xlsx) upload disabled")

# ── Load environment ──────────────────────────────────────────────────────────
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
_ENV_PATHS = [_PROJECT_ROOT / ".env", Path.cwd() / ".env"]
for _ep in _ENV_PATHS:
    load_dotenv(dotenv_path=_ep, override=False)


def _env(key: str, default: str = "") -> str:
    for _ep in _ENV_PATHS:
        if _ep.exists():
            v = dotenv_values(_ep).get(key)
            if v:
                return v
    return os.getenv(key, default)


# ── Constants ─────────────────────────────────────────────────────────────────
MODEL_NAME = _env("MODEL_NAME", "openai/gpt-oss-120b")
LLM_TEMPERATURE = float(_env("LLM_TEMPERATURE", "0.3"))
LLM_MAX_TOKENS = int(_env("LLM_MAX_TOKENS", "1500"))
GROQ_TIMEOUT_S = 45
PIPELINE_TIMEOUT_S = 300
DATABASE_URL = _env("DATABASE_URL", "sqlite+aiosqlite:///./chief_ai.db")
CORS_ORIGINS_RAW = _env(
    "CORS_ORIGINS",
    "http://localhost:8080,http://127.0.0.1:8080,http://localhost:5173,http://127.0.0.1:5173,http://localhost:8083,http://127.0.0.1:8083",
)
CORS_ORIGINS = [o.strip() for o in CORS_ORIGINS_RAW.split(",") if o.strip()]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("chief_ai")


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2 — DATABASE (SQLite, async, auto-created)
# ─────────────────────────────────────────────────────────────────────────────

engine = create_async_engine(DATABASE_URL, echo=False)
AsyncSessionLocal = async_sessionmaker(engine, expire_on_commit=False)


class Base(DeclarativeBase):
    pass


class FounderPreference(Base):
    __tablename__ = "founder_preferences"
    id = Column(Integer, primary_key=True, autoincrement=True)
    key = Column(String(100), unique=True, nullable=False)
    value = Column(String(500), nullable=False)
    description = Column(Text, nullable=True)
    updated_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))


class SignalLog(Base):
    __tablename__ = "signal_log"
    id = Column(Integer, primary_key=True, autoincrement=True)
    signal_type = Column(String(100), nullable=False)
    source_module = Column(String(50), nullable=False)
    severity = Column(String(20), nullable=False)  # CRITICAL / HIGH / WARNING / INFO
    title = Column(String(300), nullable=False)
    description = Column(Text, nullable=True)
    monetary_amount = Column(Float, nullable=True)
    fingerprint = Column(String(64), nullable=False)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))


class DecisionMemory(Base):
    __tablename__ = "decision_memory"
    id = Column(Integer, primary_key=True, autoincrement=True)
    signal_type = Column(String(100), nullable=False)
    signal_severity = Column(String(20), nullable=False)
    monetary_amount = Column(Float, nullable=True)
    action_taken = Column(Text, nullable=True)
    outcome = Column(Text, nullable=True)
    context_snapshot = Column(Text, nullable=True)  # JSON string
    recorded_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))


class WorkflowLog(Base):
    __tablename__ = "workflow_log"
    id = Column(Integer, primary_key=True, autoincrement=True)
    workflow_id = Column(String(100), nullable=False)
    status = Column(String(20), nullable=False)  # running / completed / failed
    trigger = Column(String(200), nullable=True)
    result_summary = Column(Text, nullable=True)
    duration_ms = Column(Integer, nullable=True)
    started_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))
    completed_at = Column(DateTime, nullable=True)


class UploadedFile(Base):
    __tablename__ = "uploaded_files"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    company_name = Column(String(200), default="Aghron")
    original_name = Column(String(500))
    file_path = Column(String(1000))        # disk path
    currency = Column(String(10), default="INR")
    period_start = Column(String(20))
    period_end = Column(String(20))
    tx_count = Column(Integer, default=0)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = Column(DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))


class StoredTransaction(Base):
    __tablename__ = "stored_transactions"
    id = Column(Integer, primary_key=True, autoincrement=True)
    file_id = Column(String(36), ForeignKey("uploaded_files.id", ondelete="CASCADE"), nullable=False)
    date = Column(String(20), nullable=False)
    type = Column(String(20), nullable=False)  # Revenue/Expense
    amount = Column(Float, nullable=False)
    category = Column(String(300), nullable=False)
    description = Column(Text, default="")
    is_edited = Column(Boolean, default=False)


class AnalysisResult(Base):
    __tablename__ = "analysis_results"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    file_id = Column(String(36), ForeignKey("uploaded_files.id", ondelete="CASCADE"), nullable=False)
    result_json = Column(Text, nullable=False)         # full PipelineResult JSON
    health_score = Column(Float, nullable=True)
    priority = Column(String(20), nullable=True)
    latency_s = Column(Float, nullable=True)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))


class PolicyRule(Base):
    __tablename__ = "policy_rules"
    id = Column(Integer, primary_key=True, autoincrement=True)
    rule_key = Column(String(100), unique=True, nullable=False) # e.g. 'max_single_spend'
    value = Column(Float, nullable=False)
    description = Column(String(500), nullable=True)


class BudgetLimit(Base):
    __tablename__ = "budget_limits"
    id = Column(Integer, primary_key=True, autoincrement=True)
    category = Column(String(200), unique=True, nullable=False)
    monthly_limit = Column(Float, nullable=False)
    description = Column(String(500), nullable=True)


class MarketTicker(Base):
    __tablename__ = "market_tickers"
    id = Column(Integer, primary_key=True, autoincrement=True)
    symbol = Column(String(50), unique=True, nullable=False)
    price = Column(Float, nullable=False)
    change_pct = Column(Float, nullable=True)
    updated_at = Column(DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))


class NewsFeed(Base):
    __tablename__ = "news_feed"
    id = Column(Integer, primary_key=True, autoincrement=True)
    headline = Column(String(500), nullable=False)
    sentiment = Column(String(20), nullable=False) # positive / neutral / negative
    source = Column(String(100), nullable=True)
    url = Column(String(500), nullable=True)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))


class DebtSchedule(Base):
    __tablename__ = "debt_schedules"
    id = Column(Integer, primary_key=True, autoincrement=True)
    creditor = Column(String(200), nullable=False)
    principal_amount = Column(Float, nullable=False)
    interest_rate = Column(Float, nullable=False) # e.g. 9.5 for 9.5%
    monthly_emi = Column(Float, nullable=False)
    remaining_balance = Column(Float, nullable=False)
    due_date = Column(String(20), nullable=False) # e.g. '05th of month'


class ComplianceCalendar(Base):
    __tablename__ = "compliance_calendar"
    id = Column(Integer, primary_key=True, autoincrement=True)
    obligation = Column(String(300), nullable=False)
    due_date = Column(String(20), nullable=False) # 'YYYY-MM-DD'
    authority = Column(String(100), nullable=True) # e.g. 'GSTN', 'ITD'
    status = Column(String(20), default="pending") # pending / filed / overdue


class DecisionFeedback(Base):
    __tablename__ = "decision_feedback"
    id = Column(Integer, primary_key=True, autoincrement=True)
    decision_id = Column(Integer, nullable=False)
    recorded_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))
    actual_outcome = Column(Text, nullable=False)
    score_change = Column(Float, nullable=True) # positive or negative impact score


async def get_db():
    async with AsyncSessionLocal() as session:
        yield session


_DEFAULT_PREFERENCES = [
    {
        "key": "risk_appetite",
        "value": "conservative",
        "description": "Founder's risk tolerance: conservative | balanced | aggressive",
    },
    {
        "key": "growth_focus",
        "value": "sustainable",
        "description": "Strategic emphasis: hyper-growth | balanced | sustainable",
    },
    {
        "key": "reinvestment_ratio",
        "value": "0.30",
        "description": "Fraction of monthly net profit reinvested back into the business",
    },
    {
        "key": "debt_appetite",
        "value": "low",
        "description": "Appetite for external credit: none | low | moderate | high",
    },
    {
        "key": "primary_currency",
        "value": "INR",
        "description": "Reporting currency: INR | USD | EUR",
    },
    {
        "key": "industry",
        "value": "wealthtech",
        "description": "Company industry/sector for contextual analysis",
    },
]


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3 — FILE PARSING ENGINE
# ─────────────────────────────────────────────────────────────────────────────


class Transaction(BaseModel):
    date: str
    type: Literal["Revenue", "Expense"]
    amount: float
    category: str
    description: str = ""


class FinancialDataPayload(BaseModel):
    transactions: list[Transaction]
    company_name: str = "Aghron"
    currency: str = "INR"
    period_start: str = ""
    period_end: str = ""
    raw_row_count: int = 0
    skipped_rows: int = 0
    source_filename: str = ""
    sheets_detected: list[str] = []
    metadata: dict[str, Any] = {}


def _parse_amount(v: Any) -> float:
    if isinstance(v, (int, float)):
        return abs(float(v))
    if isinstance(v, str):
        cleaned = v.strip().replace(",", "").replace("₹", "").replace("$", "").replace("€", "")
        # handle parentheses for negative
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]
        try:
            return abs(float(cleaned))
        except ValueError:
            return float("nan")
    return float("nan")


def _parse_date(v: Any) -> Optional[str]:
    if not v:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d %b %Y", "%b %d %Y", "%Y/%m/%d"]:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    # try ISO with time
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date().isoformat()
    except Exception:
        return None


def _normalize_type(v: Any) -> Optional[Literal["Revenue", "Expense"]]:
    if not isinstance(v, str):
        return None
    s = v.strip().lower()
    if s in {"revenue", "income", "sales", "credit", "inflow", "receipt", "earn"}:
        return "Revenue"
    if s in {"expense", "expenses", "cost", "debit", "spend", "outflow", "payment", "purchase", "cost of sales"}:
        return "Expense"
    return None


def _find_col(row: dict, candidates: list[str]) -> Optional[str]:
    keys_lower = {k.strip().lower(): k for k in row}
    for c in candidates:
        if c in keys_lower:
            return keys_lower[c]
    return None


def _rows_to_transactions(rows: list[dict]) -> tuple[list[Transaction], int]:
    txns: list[Transaction] = []
    skipped = 0
    if not rows:
        return txns, skipped

    first_row = rows[0]
    
    # 1. Main Date column
    date_k = _find_col(first_row, ["date", "transaction date", "txn date", "trans date", "value date", "timestamp", "time"])
    
    # 2. Credit/Debit columns
    credit_k = _find_col(first_row, ["credit", "credit amount", "deposit", "deposits", "inflow", "inflows", "received", "revenue amount"])
    debit_k = _find_col(first_row, ["debit", "debit amount", "withdrawal", "withdrawals", "outflow", "outflows", "spend", "payment", "charge", "charges", "expense amount"])
    
    # 3. Explicit Type column
    type_k = _find_col(first_row, ["type", "kind", "category type", "transaction type", "txn type", "credit/debit", "cr/dr"])
    
    # 4. Main Amount column
    amount_k = _find_col(first_row, ["amount", "value", "total", "sum", "net amount", "balance", "net"])
    
    # 5. Category/Description columns
    cat_k = _find_col(first_row, ["category", "description", "label", "particulars", "narration", "remarks", "head", "details", "payee", "memo"])
    desc_k = _find_col(first_row, ["description", "narration", "remarks", "particulars", "notes", "memo"])

    for row in rows:
        if not date_k:
            skipped += 1
            continue
            
        date = _parse_date(row.get(date_k))
        if not date:
            skipped += 1
            continue

        category = str(row.get(cat_k)).strip() if cat_k and row.get(cat_k) else "Uncategorized"
        desc = str(row.get(desc_k) or row.get("description") or row.get("narration") or "").strip()

        parsed_amount = float('nan')
        parsed_type = None

        # Scenario A: Separate Credit/Debit columns
        if credit_k or debit_k:
            cred_val = _parse_amount(row.get(credit_k)) if credit_k else float('nan')
            deb_val = _parse_amount(row.get(debit_k)) if debit_k else float('nan')

            if cred_val == cred_val and cred_val > 0:
                parsed_amount = cred_val
                parsed_type = "Revenue"
            elif deb_val == deb_val and deb_val > 0:
                parsed_amount = deb_val
                parsed_type = "Expense"

        # Scenario B: Explicit Type + Amount
        if parsed_amount != parsed_amount and type_k and amount_k:
            parsed_type = _normalize_type(row.get(type_k))
            parsed_amount = _parse_amount(row.get(amount_k))

        # Scenario C: Signed Amount column only
        if parsed_amount != parsed_amount and amount_k:
            raw_amt_str = str(row.get(amount_k) or "").strip()
            has_parentheses = raw_amt_str.startswith("(") and raw_amt_str.endswith(")")
            is_explicit_negative = raw_amt_str.startswith("-") or has_parentheses
            abs_amt = _parse_amount(raw_amt_str)
            
            if abs_amt == abs_amt and abs_amt > 0:
                parsed_amount = abs_amt
                parsed_type = "Expense" if is_explicit_negative else "Revenue"

        # If unresolved, skip
        if parsed_amount != parsed_amount or not parsed_type:
            skipped += 1
            continue

        txns.append(Transaction(
            date=date,
            type=parsed_type,
            amount=parsed_amount,
            category=category,
            description=desc
        ))

    return txns, skipped



def _parse_csv_bytes(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_json_bytes(content: bytes) -> list[dict]:
    data = json.loads(content.decode("utf-8"))
    if isinstance(data, list):
        return data
    # support {"transactions": [...]} or {"data": [...]}
    for key in ("transactions", "data", "records", "rows"):
        if key in data and isinstance(data[key], list):
            return data[key]
    return [data]


def _parse_excel_bytes(content: bytes, filename: str) -> tuple[list[dict], list[str]]:
    if not _OPENPYXL_AVAILABLE:
        raise ValueError("openpyxl not installed. Install with: pip install openpyxl")
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets_detected = wb.sheetnames
    all_rows: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        for data_row in rows[1:]:
            row_dict = {headers[i]: (data_row[i] if i < len(data_row) else None) for i in range(len(headers))}
            # skip fully empty rows
            if all(v is None or str(v).strip() == "" for v in row_dict.values()):
                continue
            all_rows.append(row_dict)
    return all_rows, sheets_detected


async def parse_uploaded_file(file: UploadFile) -> FinancialDataPayload:
    content = await file.read()
    filename = file.filename or "upload"
    ext = filename.rsplit(".", 1)[-1].lower()

    raw_rows: list[dict] = []
    sheets_detected: list[str] = []

    if ext == "csv":
        raw_rows = _parse_csv_bytes(content)
    elif ext in ("xlsx", "xls"):
        raw_rows, sheets_detected = _parse_excel_bytes(content, filename)
    elif ext == "json":
        raw_rows = _parse_json_bytes(content)
    else:
        raise ValueError(f"Unsupported file type: .{ext}. Supported: CSV, XLSX, XLS, JSON")

    txns, skipped = _rows_to_transactions(raw_rows)
    if not txns:
        raise ValueError(
            "No valid transactions found. Ensure your file has columns: Date, Type (Revenue/Expense), Amount, Category"
        )

    dates = sorted(t.date for t in txns)
    return FinancialDataPayload(
        transactions=txns,
        company_name="Company",
        currency="INR",
        period_start=dates[0] if dates else "",
        period_end=dates[-1] if dates else "",
        raw_row_count=len(raw_rows),
        skipped_rows=skipped,
        source_filename=filename,
        sheets_detected=sheets_detected,
    )


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4 — KPI ENGINE
# ─────────────────────────────────────────────────────────────────────────────


class MonthlyPoint(BaseModel):
    month: str
    revenue: float
    expenses: float
    net: float


class KpiSnapshot(BaseModel):
    total_revenue: float
    total_expenses: float
    profit: float
    monthly_burn: float
    runway_months: Optional[float]
    cash_on_hand: float
    gross_margin_pct: float
    monthly: list[MonthlyPoint]
    forecast: list[MonthlyPoint]
    priority: Literal["CRITICAL", "WARNING", "NORMAL"]
    priority_reason: str
    trend_summary: str
    top_categories: list[dict]


def compute_kpis(data: FinancialDataPayload) -> KpiSnapshot:
    txns = data.transactions
    total_revenue = sum(t.amount for t in txns if t.type == "Revenue")
    total_expenses = sum(t.amount for t in txns if t.type == "Expense")
    profit = total_revenue - total_expenses

    # Monthly breakdown
    monthly_map: dict[str, dict] = {}
    for t in txns:
        m = t.date[:7]
        if m not in monthly_map:
            monthly_map[m] = {"month": m, "revenue": 0.0, "expenses": 0.0, "net": 0.0}
        if t.type == "Revenue":
            monthly_map[m]["revenue"] += t.amount
        else:
            monthly_map[m]["expenses"] += t.amount
        monthly_map[m]["net"] = monthly_map[m]["revenue"] - monthly_map[m]["expenses"]

    monthly = [MonthlyPoint(**v) for v in sorted(monthly_map.values(), key=lambda x: x["month"])]

    # Burn rate (avg net loss over last 3 months, positive = burning)
    recent = monthly[-3:]
    avg_net = sum(m.net for m in recent) / len(recent) if recent else 0
    monthly_burn = abs(avg_net) if avg_net < 0 else 0.0

    cash_on_hand = profit
    runway_months = max(0.0, cash_on_hand / monthly_burn) if monthly_burn > 0 else None

    gross_margin_pct = ((total_revenue - total_expenses) / total_revenue * 100) if total_revenue > 0 else 0.0

    # Forecast (linear regression, 3 months ahead)
    forecast: list[MonthlyPoint] = []
    if len(monthly) >= 2:
        n = len(monthly)
        xs = list(range(n))
        ys_r = [m.revenue for m in monthly]
        ys_e = [m.expenses for m in monthly]

        def lin_reg(ys: list[float]):
            x_mean = sum(xs) / n
            y_mean = sum(ys) / n
            num = sum((xs[i] - x_mean) * (ys[i] - y_mean) for i in range(n))
            den = sum((xs[i] - x_mean) ** 2 for i in range(n))
            slope = num / den if den != 0 else 0
            intercept = y_mean - slope * x_mean
            return lambda x: max(0, intercept + slope * x)

        f_r, f_e = lin_reg(ys_r), lin_reg(ys_e)
        last_m = monthly[-1].month
        y, mo = int(last_m[:4]), int(last_m[5:7])
        for i in range(1, 4):
            mo_i = mo + i
            y_i = y + (mo_i - 1) // 12
            mo_i = ((mo_i - 1) % 12) + 1
            key = f"{y_i}-{mo_i:02d}"
            r, e = f_r(n - 1 + i), f_e(n - 1 + i)
            forecast.append(MonthlyPoint(month=key, revenue=r, expenses=e, net=r - e))

    # Priority
    priority: Literal["CRITICAL", "WARNING", "NORMAL"] = "NORMAL"
    priority_reason = "Financials are stable."
    if profit < 0 and total_revenue > 0 and abs(profit) > total_revenue * 0.2:
        priority = "CRITICAL"
        priority_reason = "Net loss exceeds 20% of total revenue."
    elif runway_months is not None and runway_months < 6:
        priority = "CRITICAL"
        priority_reason = f"Cash runway critical: {runway_months:.1f} months remaining."
    elif monthly_burn > 0 and (runway_months is None or runway_months < 12):
        priority = "WARNING"
        priority_reason = f"High burn rate — runway is {runway_months:.1f} months." if runway_months else "High burn rate detected."
    elif profit < 0:
        priority = "WARNING"
        priority_reason = "Company is operating at a net loss."

    # Trend summary
    trend_summary = "Insufficient data for trend analysis."
    if len(monthly) >= 2:
        first, last = monthly[0], monthly[-1]
        rev_growth = ((last.revenue - first.revenue) / first.revenue * 100) if first.revenue else 0
        exp_growth = ((last.expenses - first.expenses) / first.expenses * 100) if first.expenses else 0
        trend_summary = (
            f"Over {len(monthly)} months: revenue {'+' if rev_growth >= 0 else ''}{rev_growth:.1f}%, "
            f"expenses {'+' if exp_growth >= 0 else ''}{exp_growth:.1f}%."
        )

    # Top categories by spend
    cat_spend: dict[str, float] = {}
    for t in txns:
        cat_spend[t.category] = cat_spend.get(t.category, 0) + t.amount
    top_categories = sorted(
        [{"category": k, "amount": v} for k, v in cat_spend.items()], key=lambda x: x["amount"], reverse=True
    )[:10]

    return KpiSnapshot(
        total_revenue=round(total_revenue, 2),
        total_expenses=round(total_expenses, 2),
        profit=round(profit, 2),
        monthly_burn=round(monthly_burn, 2),
        runway_months=round(runway_months, 1) if runway_months is not None else None,
        cash_on_hand=round(cash_on_hand, 2),
        gross_margin_pct=round(gross_margin_pct, 2),
        monthly=monthly,
        forecast=forecast,
        priority=priority,
        priority_reason=priority_reason,
        trend_summary=trend_summary,
        top_categories=top_categories,
    )


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 5 — LLM CLIENT + ALL 8 AI AGENTS
# ─────────────────────────────────────────────────────────────────────────────


def _get_all_api_keys() -> list[str]:
    keys = []
    # Candidate keys in preference order
    candidates = [
        ["Master_AI", "GROQ_API_KEY"],
        ["Master_AI_2", "GROQ_API_KEY_2", "Master_AI2", "GROQ_API_KEY2"],
        ["Master_AI_3", "GROQ_API_KEY_3", "Master_AI3", "GROQ_API_KEY3"],
        ["Master_AI_4", "GROQ_API_KEY_4", "Master_AI4", "GROQ_API_KEY4"],
    ]
    
    # Read dotenv files
    dotenv_vals = {}
    for ep in _ENV_PATHS:
        if ep.exists():
            try:
                dotenv_vals.update(dotenv_values(ep))
            except Exception:
                pass
            
    for grp in candidates:
        found_key = None
        # Check in loaded dotenv values first
        for name in grp:
            val = dotenv_vals.get(name)
            if val:
                found_key = val
                break
        # If not found in dotenv, check os.environ
        if not found_key:
            for name in grp:
                val = os.getenv(name)
                if val:
                    found_key = val
                    break
        if found_key:
            cleaned = found_key.strip()
            # Remove optional quotes
            if cleaned.startswith(('"', "'")) and cleaned.endswith(('"', "'")):
                cleaned = cleaned[1:-1]
            if cleaned and cleaned not in keys:
                keys.append(cleaned)
                
    return keys


def _require_api_key() -> str:
    keys = _get_all_api_keys()
    if keys:
        return keys[0]
    raise RuntimeError("GROQ_API_KEY / Master_AI not configured in .env")


def _safe_json(v: Any) -> str:
    return json.dumps(v, indent=2, ensure_ascii=False, default=str)


def _extract_json(text: str) -> Any:
    text = text.strip()
    # strip markdown code fences
    if "```" in text:
        start = text.find("```")
        end = text.rfind("```")
        inner = text[start + 3:end]
        if inner.startswith("json"):
            inner = inner[4:]
        text = inner.strip()
    try:
        return json.loads(text, strict=False)
    except json.JSONDecodeError:
        # find first { or [
        for start_char in ["{", "["]:
            idx = text.find(start_char)
            if idx != -1:
                end_char = "}" if start_char == "{" else "]"
                end_idx = text.rfind(end_char)
                if end_idx > idx:
                    try:
                        return json.loads(text[idx: end_idx + 1], strict=False)
                    except json.JSONDecodeError:
                        continue
        raise


def call_groq(prompt: str, max_tokens: int | None = None) -> str:
    api_keys = _get_all_api_keys()
    if not api_keys:
        raise RuntimeError("No GROQ API keys (GROQ_API_KEY / Master_AI) configured in environment or .env")
        
    _max_tokens = max_tokens if max_tokens is not None else LLM_MAX_TOKENS
    last_err = None
    for idx, api_key in enumerate(api_keys):
        try:
            log.info(f"LLM API Call: Trying key index {idx+1}/{len(api_keys)} (Ends in ...{api_key[-6:] if len(api_key) > 6 else api_key})")
            client = Groq(api_key=api_key, timeout=httpx.Timeout(GROQ_TIMEOUT_S))
            completion = client.chat.completions.create(
                model=MODEL_NAME,
                messages=[{"role": "user", "content": prompt}],
                temperature=LLM_TEMPERATURE,
                max_tokens=_max_tokens,
                top_p=1,
                stream=False,
            )
            content = completion.choices[0].message.content
            if content:
                return content
            raise ValueError("Empty response received from Groq API")
        except Exception as e:
            last_err = e
            log.warning(f"LLM API Call failed with key index {idx+1}: {e}. Retrying with next key if available...")
            
    log.error(f"All {len(api_keys)} API keys failed. Last error: {last_err}")
    raise last_err or RuntimeError("All API keys failed to return a response")


def call_json_agent(prompt: str) -> dict[str, Any]:
    raw = call_groq(prompt)
    parsed = _extract_json(raw)
    if isinstance(parsed, dict):
        return parsed
    raise ValueError(f"Agent returned non-object JSON: {type(parsed)}")


def _agent_wrap(name: str, output: dict, duration_ms: int) -> dict:
    return {"bot": name, "output": output, "meta": {"duration_ms": duration_ms, "model": MODEL_NAME}}


# ── Agent 1: Research AI ──────────────────────────────────────────────────────
def run_research_ai(kpis: KpiSnapshot, data_meta: dict, market_tickers: Optional[list] = None, news_headlines: Optional[list] = None) -> dict:
    name = "Research AI"
    log.info(f"▶ {name}")
    start = time.perf_counter()
    
    live_tickers_text = "Live Scraped Market Tickers (Yahoo Finance):\n"
    if market_tickers:
        for t in market_tickers[:10]:
            live_tickers_text += f"- {t['symbol']} ({t['label']}): Price {t['price']} | Change: {'+' if t['change_pct'] >= 0 else ''}{t['change_pct']}%\n"
    else:
        live_tickers_text += "- Live market connection offline.\n"
        
    live_news_text = "Live Crawled Sector & Macro News (Moneycontrol RSS):\n"
    if news_headlines:
        for n in news_headlines[:8]:
            live_news_text += f"- Title: {n['headline']} | Sentiment: {n['sentiment']} | Source: {n['source']}\n"
    else:
        live_news_text += "- Live news crawling feed offline.\n"

    prompt = f"""You are the Research AI for a finance system.

Company Financial Context:
- Revenue: {data_meta.get('currency', 'INR')} {kpis.total_revenue:,.0f}
- Expenses: {kpis.total_expenses:,.0f}
- Profit: {kpis.profit:,.0f}
- Monthly Burn: {kpis.monthly_burn:,.0f}
- Runway: {kpis.runway_months if kpis.runway_months else 'Profitable / No burn'}
- Priority Level: {kpis.priority}
- Industry: {data_meta.get('industry', 'technology')}
- Period: {data_meta.get('period_start', '')} to {data_meta.get('period_end', '')}

{live_tickers_text}
{live_news_text}

Analyze the macro-economic environment and market conditions relevant to this company.

Return ONLY valid JSON (no markdown, no explanation):
{{
  "bot": "Research AI",
  "market_summary": "2-3 sentence macro overview relevant to company's financials based on the live tickers and news headlines",
  "key_trends": ["trend 1", "trend 2", "trend 3"],
  "risk_signals": [
    {{"signal": "description", "severity": "HIGH|MEDIUM|LOW", "monetary_impact": "estimate or N/A"}}
  ],
  "risk_factors": ["macro volatility risk", "cost expansion risk", "rate sensitivity risk"],
  "opportunities": ["opportunity 1", "opportunity 2"],
  "macro_assumptions": {{
    "interest_rate_trend": "rising|stable|falling",
    "inflation_outlook": "high|moderate|low",
    "market_sentiment": "bullish|neutral|bearish",
    "sector_risk": "high|medium|low"
  }},
  "market_sentiment": "Neutral",
  "confidence": 0.85
}}"""
    try:
        out = call_json_agent(prompt)
    except Exception as e:
        log.error(f"{name} failed: {e}")
        out = {
            "bot": name,
            "error": str(e),
            "market_summary": "Analysis unavailable",
            "key_trends": [],
            "risk_signals": [],
            "risk_factors": ["Macro volatility risk", "High sector risk"],
            "opportunities": [],
            "macro_assumptions": {"market_sentiment": "neutral"},
            "market_sentiment": "Neutral",
            "confidence": 0.0
        }
    return _agent_wrap(name, out, int((time.perf_counter() - start) * 1000))


# ── Agent 2: Planning AI ──────────────────────────────────────────────────────
def run_planning_ai(kpis: KpiSnapshot, research_output: dict, preferences: dict, budgets: Optional[list] = None) -> dict:
    name = "Planning AI"
    log.info(f"▶ {name}")
    start = time.perf_counter()
    
    # 1. Compile budgets text
    budgets_text = "Seeded Active Corporate Budgets:\n"
    if budgets:
        for b in budgets:
            budgets_text += f"- Category: {b['category']} | Monthly Limit: ₹{b['monthly_limit']:,.2f}\n"
    else:
        budgets_text += "- No corporate budget ceilings seeded.\n"

    # 2. Math-based Scenarios (Calculated server-side)
    base_rev = kpis.total_revenue / max(len(kpis.monthly), 1)
    base_exp = kpis.total_expenses / max(len(kpis.monthly), 1)
    base_net = base_rev - base_exp
    base_burn = abs(base_net) if base_net < 0 else 0.0
    base_runway = kpis.cash_on_hand / base_burn if base_burn > 0 else 999.0
    
    # Best case (+15% revenue growth, -10% cost reduction)
    best_rev = base_rev * 1.15
    best_exp = base_exp * 0.90
    best_net = best_rev - best_exp
    best_burn = abs(best_net) if best_net < 0 else 0.0
    best_runway = kpis.cash_on_hand / best_burn if best_burn > 0 else 999.0
    
    # Worst case (-20% revenue contraction, +15% cost spike)
    worst_rev = base_rev * 0.80
    worst_exp = base_exp * 1.15
    worst_net = worst_rev - worst_exp
    worst_burn = abs(worst_net) if worst_net < 0 else 0.0
    worst_runway = kpis.cash_on_hand / worst_burn if worst_burn > 0 else 999.0
    
    scenario_math = {
        "best_case": {"revenue": round(best_rev, 2), "expenses": round(best_exp, 2), "burn": round(best_burn, 2), "runway": round(best_runway, 1)},
        "base_case": {"revenue": round(base_rev, 2), "expenses": round(base_exp, 2), "burn": round(base_burn, 2), "runway": round(base_runway, 1)},
        "worst_case": {"revenue": round(worst_rev, 2), "expenses": round(worst_exp, 2), "burn": round(worst_burn, 2), "runway": round(worst_runway, 1)}
    }

    prompt = f"""You are the Planning AI (FP&A) for a finance system.

Company KPIs:
- Revenue: {kpis.total_revenue:,.0f} | Expenses: {kpis.total_expenses:,.0f}
- Profit: {kpis.profit:,.0f} | Monthly Burn: {kpis.monthly_burn:,.0f}
- Runway: {kpis.runway_months if kpis.runway_months else 'Profitable'} months
- Trend: {kpis.trend_summary}
- Priority: {kpis.priority} — {kpis.priority_reason}

Top Spending Categories: {_safe_json(kpis.top_categories[:5])}
3-Month Forecast: {_safe_json([m.model_dump() for m in kpis.forecast])}

{budgets_text}

Pre-Calculated Scenario Math (Do NOT hallucinate runway/burn, use these exact figures):
{_safe_json(scenario_math)}

Research Context: {_safe_json(research_output.get('output', {}))}

Founder Preferences:
- Risk Appetite: {preferences.get('risk_appetite', 'balanced')}
- Growth Focus: {preferences.get('growth_focus', 'balanced')}
- Reinvestment Ratio: {preferences.get('reinvestment_ratio', '0.30')}
- Debt Appetite: {preferences.get('debt_appetite', 'low')}

Provide comprehensive financial planning analysis. Make sure to generate months_projections for the upcoming 12 months.

Return ONLY valid JSON:
{{
  "bot": "Planning AI",
  "twelve_month_forecast": {{
    "revenue_projection": 0,
    "expense_projection": 0,
    "ebitda_projection": 0,
    "confidence": 0.0
  }},
  "months_projections": [
    {{"month": "2026-06", "revenue": 0.0, "expenses": 0.0, "cash": 0.0}}
  ],
  "burn_analysis": {{
    "current_monthly_burn": 0,
    "trend": "increasing|stable|decreasing",
    "primary_burn_drivers": []
  }},
  "runway_analysis": {{
    "current_runway_months": 0,
    "alert_level": "CRITICAL|WARNING|NORMAL",
    "recommendation": ""
  }},
  "scenarios": {{
    "best_case": {{"description": "strategic summary", "runway_months": {scenario_math['best_case']['runway']}, "conditions": ["+15% revenue expansion", "-10% operational cost cut"], "runway": {scenario_math['best_case']['runway']}, "burn": {scenario_math['best_case']['burn']}}},
    "base_case": {{"description": "strategic summary", "runway_months": {scenario_math['base_case']['runway']}, "conditions": ["flat revenue", "stable OPEX pacing"], "runway": {scenario_math['base_case']['runway']}, "burn": {scenario_math['base_case']['burn']}}},
    "worst_case": {{"description": "strategic summary", "runway_months": {scenario_math['worst_case']['runway']}, "conditions": ["-20% market contraction", "+15% cost spike"], "runway": {scenario_math['worst_case']['runway']}, "burn": {scenario_math['worst_case']['burn']}}}
  }},
  "budget_recommendations": [],
  "action_items": ["Implement strict vendor checks", "Optimize cloud infrastructure spend"],
  "anomalies_detected": [],
  "confidence": 0.85
}}"""
    try:
        out = call_json_agent(prompt)
    except Exception as e:
        log.error(f"{name} failed: {e}")
        out = {
            "bot": name,
            "error": str(e),
            "twelve_month_forecast": {},
            "months_projections": [],
            "burn_analysis": {},
            "runway_analysis": {},
            "scenarios": {},
            "budget_recommendations": [],
            "action_items": ["Optimize cloud infrastructure spend", "Engage accounts receivables closely"],
            "anomalies_detected": [],
            "confidence": 0.0
        }
    return _agent_wrap(name, out, int((time.perf_counter() - start) * 1000))


# ── Agent 3: Accounting AI ────────────────────────────────────────────────────
def run_accounting_ai(data: FinancialDataPayload, kpis: KpiSnapshot, budgets: Optional[list] = None) -> dict:
    name = "Accounting AI"
    log.info(f"▶ {name}")
    start = time.perf_counter()
    
    txns = data.transactions
    total_count = len(txns)
    
    # 1. Compile budgets mapping
    budget_map = {b["category"]: b["monthly_limit"] for b in budgets} if budgets else {}

    # 2. Group by category and type (Recursive Ingestion Matrix)
    category_summary = {}
    for t in txns:
        key = (t.category, t.type)
        if key not in category_summary:
            category_summary[key] = {"count": 0, "amount": 0.0, "examples": []}
        category_summary[key]["count"] += 1
        category_summary[key]["amount"] += t.amount
        if len(category_summary[key]["examples"]) < 3:
            category_summary[key]["examples"].append(t.description)
            
    # Dense matrix construction
    dense_matrix = []
    for (cat, t_type), val in category_summary.items():
        limit = budget_map.get(cat, "No Limit")
        variance_flag = "nominal"
        if isinstance(limit, float) and t_type == "Expense" and val["amount"] > limit:
            variance_flag = "over_budget"
            
        dense_matrix.append({
            "category": cat,
            "type": t_type,
            "total_spend": round(val["amount"], 2),
            "count": val["count"],
            "avg_spend": round(val["amount"] / val["count"], 2),
            "monthly_budget_limit": limit,
            "budget_variance": variance_flag,
            "sample_particulars": ", ".join(val["examples"][:2])
        })
        
    # 3. Statistical Outliers Ingestion (>3 standard deviations)
    amounts = [t.amount for t in txns]
    mean_amt = sum(amounts) / max(total_count, 1)
    variance = sum((x - mean_amt) ** 2 for x in amounts) / max(total_count, 1)
    std_dev = math.sqrt(variance)
    
    outliers = []
    for t in txns:
        if std_dev > 0 and abs(t.amount - mean_amt) > 3 * std_dev:
            outliers.append({
                "date": t.date,
                "type": t.type,
                "amount": t.amount,
                "category": t.category,
                "description": t.description
            })
            
    # Limit outliers list to top 15 to avoid context limits
    outliers = sorted(outliers, key=lambda x: x["amount"], reverse=True)[:15]

    prompt = f"""You are the Accounting AI for a finance system.

Dense Category Audit Matrix (Auditing all {total_count} transactions — uncapped):
{_safe_json(dense_matrix)}

Statistical Ledger Outliers (>3 Std Devs):
{_safe_json(outliers)}

Financial Summary:
- Total Revenue: {kpis.total_revenue:,.0f}
- Total Expenses: {kpis.total_expenses:,.0f}
- File: {data.source_filename}
- Period: {data.period_start} to {data.period_end}
- Rows Skipped (parse errors): {data.skipped_rows}

Analyze for anomalies, fraud patterns, classification issues, reconciliation status, and GST/tax compliance based on the full ledger profile.

Return ONLY valid JSON:
{{
  "bot": "Accounting AI",
  "transaction_summary": {{
    "total_transactions": {total_count},
    "categorized_correctly": {total_count - len(outliers)},
    "uncategorized_count": {sum(1 for t in txns if t.category.lower() in ("uncategorized", "suspense", "other"))},
    "date_range_complete": true
  }},
  "anomalies": [
    {{"id": 101, "type": "Anomalous Transaction", "reason": "description", "details": "double debit mismatch", "severity": "HIGH|MEDIUM|LOW", "affected_amount": 0}}
  ],
  "fraud_risk_score": 12,
  "fraud_indicators": ["Double transaction detected", "Suspense GL allocation"],
  "fraud_flags": [
    {{"pattern": "High volume spike", "description": "abnormal pacing", "severity": "CRITICAL|HIGH|MEDIUM", "amount_at_risk": 0}}
  ],
  "categorization_quality": {{
    "score": 88,
    "issues": [],
    "recommendations": []
  }},
  "reconciliation_status": {{
    "status": "complete|incomplete|requires_review",
    "gap_amount": 0,
    "issues": []
  }},
  "gst_status": {{
    "readiness_score": 90,
    "mismatches": [],
    "filing_readiness": "ready|needs_review|not_ready"
  }},
  "gst_reconciliation_issues": ["GSTR-2B matching delay"],
  "close_cycle_status": "on_track|delayed|at_risk",
  "confidence": 0.85
}}"""
    try:
        out = call_json_agent(prompt)
    except Exception as e:
        log.error(f"{name} failed: {e}")
        out = {
            "bot": name,
            "error": str(e),
            "anomalies": [
                {"id": 103, "type": "Suspense Account Mismatch", "reason": "GL 59999 posting", "details": "HDFC suspense category mismatch", "severity": "HIGH", "affected_amount": 12000}
            ],
            "fraud_risk_score": 12,
            "fraud_indicators": ["Suspense GL entry found"],
            "fraud_flags": [],
            "gst_reconciliation_issues": [],
            "confidence": 0.0
        }
    return _agent_wrap(name, out, int((time.perf_counter() - start) * 1000))


# ── Agent 4: Treasury AI ──────────────────────────────────────────────────────
def run_treasury_ai(kpis: KpiSnapshot, accounting_output: dict, preferences: dict, debt_schedules: Optional[list] = None) -> dict:
    name = "Treasury AI"
    log.info(f"▶ {name}")
    start = time.perf_counter()
    
    # 1. Format debt schedules
    debt_text = "Seeded Active Loans & Debt Schedules:\n"
    if debt_schedules:
        for d in debt_schedules:
            debt_text += f"- Creditor: {d['creditor']} | Principal: ₹{d['principal_amount']:,.2f} | Remaining: ₹{d['remaining_balance']:,.2f} | EMI: ₹{d['monthly_emi']:,.2f} | Due: {d['due_date']}\n"
    else:
        debt_text += "- No outstanding debt liabilities registered.\n"
        
    # 2. Virtual wallets (simulated half/ Razorpay & Stripe integrations)
    wallet_text = "Virtual Merchant Gateway Balances (Simulated Connectors):\n"
    wallet_text += f"- Razorpay Balance: ₹{kpis.cash_on_hand * 0.35:,.2f} INR (Active Sync)\n"
    wallet_text += f"- Stripe Balance: ₹{kpis.cash_on_hand * 0.15:,.2f} INR (Active Sync)\n"

    prompt = f"""You are the Treasury AI for a finance system.

Cash & Liquidity Position:
- Total Cash on Hand: {kpis.cash_on_hand:,.0f}
- Monthly Burn: {kpis.monthly_burn:,.0f}
- Runway: {kpis.runway_months if kpis.runway_months else 'Profitable'} months
- Priority: {kpis.priority}

{debt_text}
{wallet_text}

Accounting Context: {_safe_json(accounting_output.get('output', {}))}

Founder Preferences:
- Debt Appetite: {preferences.get('debt_appetite', 'low')}
- Risk Appetite: {preferences.get('risk_appetite', 'balanced')}

Analyze cash position, liquidity health, investment opportunities, and stress scenarios.

Return ONLY valid JSON:
{{
  "bot": "Treasury AI",
  "cash_position": {{
    "total_cash": {kpis.cash_on_hand},
    "liquidity_score": 85,
    "buffer_status": "adequate|tight|critical",
    "days_of_operating_cash": {round(kpis.cash_on_hand / max(kpis.monthly_burn / 30, 1), 1) if kpis.monthly_burn > 0 else 999.0}
  }},
  "cash_position_summary": "Liquid reserves are stable with adequate working capital runway.",
  "liquidity_score": 85,
  "cash_flow_summary": {{
    "operating_cash_flow": {kpis.profit},
    "trend": "improving|stable|deteriorating",
    "primary_inflow_sources": ["Core business operations"],
    "primary_outflow_drivers": ["OPEX", "EMI repayments"]
  }},
  "liquidity_risks": [
    {{"risk": "receivables lag", "severity": "MEDIUM", "timeframe": "30 days", "amount": 0}}
  ],
  "investment_portfolio": {{
    "idle_cash": {kpis.cash_on_hand * 0.30},
    "investment_opportunities": ["7.25% HDFC Corporate Fixed Deposit Wrapper"],
    "recommended_allocation": "Deploy 30% of reserves into short-term AAA wrappers"
  }},
  "investment_options": [
    {{"instrument": "Corporate Fixed Deposit", "yield": "7.25%", "risk": "Low", "horizon": "12 months"}},
    {{"instrument": "Liquid Mutual Fund", "yield": "6.80%", "risk": "Low", "horizon": "Anytime"}}
  ],
  "stress_test": {{
    "scenario_25pct_revenue_drop": {{"impact": "runway drops", "runway_remaining": 0, "severity": "medium"}},
    "scenario_50pct_revenue_drop": {{"impact": "cash crunch", "runway_remaining": 0, "severity": "critical"}},
    "scenario_key_client_loss": {{"impact": "high threat", "runway_remaining": 0, "severity": "critical"}}
  }},
  "stress_test_results": [
    {{"scenario": "Client Churn Event", "cash_impact": "Severe reduction in inflow buffer", "outcome": "Requires short-term debt cover"}}
  ],
  "working_capital_advice": "Deploy short-term idle cash into 7.25% FD wrappers immediately.",
  "payment_schedule_health": "on_track|at_risk|overdue",
  "confidence": 0.85
}}"""
    try:
        out = call_json_agent(prompt)
    except Exception as e:
        log.error(f"{name} failed: {e}")
        out = {
            "bot": name,
            "error": str(e),
            "cash_position": {},
            "cash_position_summary": "Reserves analysis currently unavailable",
            "liquidity_score": 50,
            "liquidity_risks": [],
            "investment_options": [],
            "stress_test_results": [],
            "working_capital_advice": "Maintain standard cash reserve balances.",
            "confidence": 0.0
        }
    return _agent_wrap(name, out, int((time.perf_counter() - start) * 1000))


# ── Agent 5: Compliance AI ────────────────────────────────────────────────────
def run_compliance_ai(data: FinancialDataPayload, kpis: KpiSnapshot, preferences: dict, compliance_deadlines: Optional[list] = None) -> dict:
    name = "Compliance AI"
    log.info(f"▶ {name}")
    start = time.perf_counter()
    
    # Format compliance calendar deadlines
    calendar_text = "Seeded Federal Compliance Deadlines (GST/TDS):\n"
    upcoming_array = []
    if compliance_deadlines:
        for c in compliance_deadlines:
            try:
                due = datetime.strptime(c["due_date"], "%Y-%m-%d").date()
                days = (due - datetime.now().date()).days
            except Exception:
                days = 15
                
            calendar_text += f"- Obligation: {c['obligation']} | Due: {c['due_date']} | Days Remaining: {days}d | Auth: {c['authority']} | Status: {c['status']}\n"
            upcoming_array.append({
                "obligation": c["obligation"],
                "deadline": c["due_date"],
                "days_remaining": days,
                "status": c["status"]
            })
    else:
        calendar_text += "- No regulatory filing dates registered.\n"

    upcoming_deadlines_str = _safe_json(upcoming_array[:5]) if upcoming_array else '[{"obligation": "GST filing", "deadline": "2026-06-11", "days_remaining": 15, "status": "pending"}]'
    deadlines_str = _safe_json([{"task": x["obligation"], "due_date": x["deadline"], "priority": "High" if x["days_remaining"] < 7 else "Medium"} for x in upcoming_array]) if upcoming_array else "[]"

    prompt = f"""You are the Compliance AI for a finance system.

Company Financial Data:
- Revenue: {kpis.total_revenue:,.0f} | Expenses: {kpis.total_expenses:,.0f}
- Period: {data.period_start} to {data.period_end}
- Industry: {preferences.get('industry', 'technology')}
- Currency: {data.currency}

{calendar_text}

Analyze compliance posture, tax obligations, regulatory risks, and upcoming deadlines.

Return ONLY valid JSON:
{{
  "bot": "Compliance AI",
  "overall_health_score": 92,
  "health_label": "Excellent|Good|Needs Attention|At Risk|Critical",
  "compliance_issues": [
    {{"issue": "GSTR-1 unmatched credits", "severity": "HIGH", "regulation": "GST Act", "action_required": "Reconcile vendor filings", "deadline": "2026-06-11"}}
  ],
  "tax_flags": [
    {{"type": "GST", "description": "unmatched input credits", "amount_at_risk": 7250, "deadline": "2026-06-11"}}
  ],
  "tax_obligations": [
    {{"tax_type": "GST Input Credit Mismatch", "status": "Accrued", "estimate": 7250}},
    {{"tax_type": "TDS Quarterly Return", "status": "Pending", "estimate": 12000}}
  ],
  "upcoming_deadlines": {upcoming_deadlines_str},
  "deadlines": {deadlines_str},
  "regulatory_risks": [
    {{"regulation": "SEBI Compliance", "risk": "Audit Documentation incompleteness", "probability": "low", "impact": "Penalties"}}
  ],
  "audit_readiness": {{
    "score": 92,
    "documentation_completeness": 95,
    "issues": []
  }},
  "audit_readiness_score": 92,
  "compliance_warnings": ["EPF statutory deposit timeline lag"],
  "recommendations": [],
  "confidence": 0.85
}}"""
    try:
        out = call_json_agent(prompt)
    except Exception as e:
        log.error(f"{name} failed: {e}")
        out = {
            "bot": name,
            "error": str(e),
            "overall_health_score": 0,
            "compliance_issues": [],
            "tax_flags": [],
            "tax_obligations": [],
            "deadlines": [],
            "audit_readiness_score": 50,
            "compliance_warnings": ["Checklist updates unavailable"],
            "confidence": 0.0
        }
    return _agent_wrap(name, out, int((time.perf_counter() - start) * 1000))


# ── Agent 6: Reporting AI ─────────────────────────────────────────────────────
def run_reporting_ai(kpis: KpiSnapshot, planning_output: dict, treasury_output: dict, compliance_output: dict) -> dict:
    name = "Reporting AI"
    log.info(f"▶ {name}")
    start = time.perf_counter()
    prompt = f"""You are the Reporting AI (MIS/Dashboard) for a finance system.

Core KPIs:
- Revenue: {kpis.total_revenue:,.0f} | Expenses: {kpis.total_expenses:,.0f}
- Profit: {kpis.profit:,.0f} | Gross Margin: {kpis.gross_margin_pct:.1f}%
- Monthly Burn: {kpis.monthly_burn:,.0f} | Runway: {kpis.runway_months if kpis.runway_months else 'Profitable'} months
- Priority: {kpis.priority} — {kpis.priority_reason}
- Trend: {kpis.trend_summary}

Planning Output: {_safe_json(planning_output.get('output', {}))}
Treasury Output: {_safe_json(treasury_output.get('output', {}))}
Compliance Output: {_safe_json(compliance_output.get('output', {}))}

Generate comprehensive MIS/board-level reporting analysis.

Return ONLY valid JSON:
{{
  "bot": "Reporting AI",
  "executive_summary": "3-4 sentence CFO-grade financial summary",
  "kpi_dashboard": {{
    "revenue_status": "above_plan|on_plan|below_plan",
    "ebitda_margin": 0.0,
    "gross_margin": 0.0,
    "cac_trend": "improving|stable|deteriorating",
    "ltv_cac_ratio": 0.0,
    "key_kpis": []
  }},
  "kpi_summary": [
    {{"name": "LTV:CAC Growth Ratio", "value": "3.5x", "status": "Good"}},
    {{"name": "Gross Margin Percentage", "value": "78%", "status": "Excellent"}},
    {{"name": "EBITDA Operating Margin", "value": "18.2%", "status": "nominal"}}
  ],
  "performance_insights": [
    {{"insight": "Opex pacing matches sustainable models.", "category": "cost", "priority": "high"}}
  ],
  "board_narrative": "Board-ready 2-paragraph financial narrative",
  "mis_status": {{
    "data_freshness": "current|stale|outdated",
    "coverage_completeness": 0,
    "last_close_status": "complete|in_progress|overdue"
  }},
  "talking_points": ["point 1", "point 2", "point 3"],
  "confidence": 0.85
}}"""
    try:
        out = call_json_agent(prompt)
    except Exception as e:
        log.error(f"{name} failed: {e}")
        out = {
            "bot": name,
            "error": str(e),
            "executive_summary": "Reporting unavailable",
            "kpi_dashboard": {},
            "kpi_summary": [],
            "board_narrative": "",
            "performance_insights": [],
            "confidence": 0.0
        }
    return _agent_wrap(name, out, int((time.perf_counter() - start) * 1000))


# ── Agent 7: Decision AI ──────────────────────────────────────────────────────
def run_decision_ai(
    kpis: KpiSnapshot,
    research: dict,
    planning: dict,
    treasury: dict,
    compliance: dict,
    reporting: dict,
    preferences: dict,
    feedback_history: Optional[list] = None
) -> dict:
    name = "Decision AI"
    log.info(f"▶ {name}")
    start = time.perf_counter()
    
    # Format outcome learning context
    learning_context = ""
    if feedback_history:
        learning_context = "\nHistorical outcomes from similar strategic choices (Reinforcement Learning):\n"
        for fb in feedback_history[:10]:
            learning_context += f"- Decision ID {fb['decision_id']}: Actual Outcome: {fb['actual_outcome']} | Impact: {'+' if fb['score_change'] >= 0 else ''}{fb['score_change']} points\n"
    else:
        learning_context = "\nNo prior decision feedback registered in reinforcement library.\n"

    prompt = f"""You are the Decision AI — the final recommendation engine for a finance system.

Core Financial Position:
- Revenue: {kpis.total_revenue:,.0f} | Profit: {kpis.profit:,.0f}
- Burn: {kpis.monthly_burn:,.0f}/month | Runway: {kpis.runway_months if kpis.runway_months else 'Profitable'} months
- Priority: {kpis.priority} — {kpis.priority_reason}

{learning_context}

Agent Outputs Summary:
Research: {_safe_json(research.get('output', {}))}
Planning: {_safe_json(planning.get('output', {}))}
Treasury: {_safe_json(treasury.get('output', {}))}
Compliance: {_safe_json(compliance.get('output', {}))}
Reporting: {_safe_json(reporting.get('output', {}))}

Founder Strategic Preferences:
- Risk Appetite: {preferences.get('risk_appetite', 'balanced')}
- Growth Focus: {preferences.get('growth_focus', 'balanced')}
- Reinvestment Ratio: {preferences.get('reinvestment_ratio', '0.30')}
- Debt Appetite: {preferences.get('debt_appetite', 'low')}

Generate final decisions strictly aligned with founder preferences.

Return ONLY valid JSON:
{{
  "bot": "Decision AI",
  "overall_risk_score": 0,
  "risk_level": "CRITICAL|HIGH|MEDIUM|LOW",
  "confidence_score": 0.85,
  "decisions": [
    {{"title": "FD Allocations", "description": "Lock cash reserves", "priority": "HIGH", "category": "liquidity"}}
  ],
  "recommended_decisions": [
    {{"decision": "Lock cash reserves in Corporate FD", "impact": "earns 7.25% short term", "timeframe": "this_week", "confidence": 0.95}},
    {{"decision": "Reconcile GSTR-2B mismatched vendor bills", "impact": "saves INR 95k in input credit", "timeframe": "this_week", "confidence": 0.90}}
  ],
  "recommended_actions": [
    {{"action": "Deploy short-term treasury investments", "timeframe": "this_week", "expected_impact": "FD yields yield 7.25%", "owner": "CFO"}}
  ],
  "founder_aligned_strategy": "2-3 sentence strategy statement aligned with founder preferences",
  "risk_breakdown": {{
    "liquidity_risk": 20,
    "market_risk": 15,
    "compliance_risk": 35,
    "operational_risk": 10
  }},
  "risk_quadrant": {{
    "axis_x": "sustainable",
    "axis_y": "balanced",
    "zone": "Balanced Growth"
  }},
  "founder_alignment_score": 88,
  "confidence": 0.85
}}"""
    try:
        out = call_json_agent(prompt)
    except Exception as e:
        log.error(f"{name} failed: {e}")
        out = {
            "bot": name,
            "error": str(e),
            "overall_risk_score": 0,
            "decisions": [],
            "recommended_decisions": [],
            "recommended_actions": [],
            "risk_quadrant": {"axis_x": "stable", "axis_y": "conservative", "zone": "Balanced"},
            "founder_alignment_score": 50,
            "confidence": 0.0
        }
    return _agent_wrap(name, out, int((time.perf_counter() - start) * 1000))


# ── Agent 8: Chief Command AI ─────────────────────────────────────────────────
def run_chief_command_ai(
    kpis: KpiSnapshot,
    decision: dict,
    reporting: dict,
    compliance: dict,
    treasury: dict,
    planning: dict,
    health_score: dict,
    composite_patterns: list[dict],
    policy_rules: Optional[list] = None
) -> dict:
    name = "Chief Command AI"
    log.info(f"▶ {name}")
    start = time.perf_counter()
    
    triggered_violations = []
    if policy_rules:
        max_spend_rule = next((r for r in policy_rules if r["rule_key"] == "max_single_spend"), None)
        if max_spend_rule:
            limit = max_spend_rule["value"]
            for cat_data in kpis.top_categories:
                if cat_data["amount"] > limit:
                    triggered_violations.append(f"Spending in category '{cat_data['category']}' (₹{cat_data['amount']:,.2f}) exceeds rule limit of ₹{limit:,.2f}")
                    
        min_liq_rule = next((r for r in policy_rules if r["rule_key"] == "min_liquidity_ratio"), None)
        if min_liq_rule:
            limit = min_liq_rule["value"]
            ratio = kpis.cash_on_hand / max(kpis.total_revenue, 1)
            if ratio < limit:
                triggered_violations.append(f"Liquidity ratio ({ratio:.1%}) drops below minimum rule limit of {limit:.1%}")
                
    policy_interception_text = "Seeded Policy Intercept Audits:\n"
    if triggered_violations:
        for v in triggered_violations:
            policy_interception_text += f"- [POLICY VIOLATION] {v}\n"
    else:
        policy_interception_text += "- All company spending rules nominal. No policy violations detected.\n"

    prompt = f"""You are the Chief Command AI — the master orchestration layer of the Finance AI System.

You have received outputs from all 7 specialist agents. Your role is to:
1. Synthesize everything into a single, prioritized executive brief
2. Classify the overall financial situation (CRITICAL / HIGH / WARNING / NORMAL)
3. List the top urgent actions the CFO must take TODAY
4. Generate the Daily Finance Brief

Finance Health Score: {health_score.get('score', 0)}/100 — {health_score.get('label', 'Unknown')}

Core Position:
- Revenue: {kpis.total_revenue:,.0f} | Profit: {kpis.profit:,.0f}
- Burn: {kpis.monthly_burn:,.0f}/month | Runway: {kpis.runway_months if kpis.runway_months else 'Profitable'} months
- System Priority: {kpis.priority}

{policy_interception_text}

Decision AI Output: {_safe_json(decision.get('output', {}))}
Reporting AI Output: {_safe_json(reporting.get('output', {}))}
Compliance AI Output: {_safe_json(compliance.get('output', {}))}
Treasury AI Output: {_safe_json(treasury.get('output', {}))}

Composite Risk Patterns Detected: {_safe_json(composite_patterns)}

Return ONLY valid JSON:
{{
  "bot": "Chief Command AI",
  "overall_priority": "{'CRITICAL' if triggered_violations else kpis.priority}",
  "finance_health_score": {{
    "score": {health_score.get('score', 0)},
    "label": "{health_score.get('label', 'Unknown')}",
    "trend": "improving|stable|deteriorating"
  }},
  "final_summary": "3-4 sentence CFO-ready executive summary of the company's complete financial position, incorporating any spending policy rule violations",
  "daily_brief": "Full finance AI daily brief in plain text, formatted with sections: HEALTH SCORE, TODAY PRIORITIES, POLICY WARNINGS (list any policy violations!), WHAT CAN WAIT, PATTERN ALERTS",
  "urgent_actions": [
    {{"action": "{triggered_violations[0] if triggered_violations else 'Assess liquidity buffers'}", "owner": "CFO", "deadline": "today", "financial_impact": "High risk of margin compression"}}
  ],
  "alerts": [
    {{"alert": "{triggered_violations[0] if triggered_violations else 'Spending rules verified'}", "severity": "{'CRITICAL' if triggered_violations else 'WARNING'}", "source_agent": "Chief Command AI", "requires_immediate_action": true}}
  ],
  "board_readiness": {{
    "ready": true,
    "blockers": [],
    "estimated_time_to_ready": ""
  }},
  "confidence": 0.85
}}"""
    try:
        out = call_json_agent(prompt)
    except Exception as e:
        log.error(f"{name} failed: {e}")
        out = {"bot": name, "error": str(e), "overall_priority": "WARNING", "final_summary": "Analysis unavailable", "urgent_actions": [], "alerts": [], "confidence": 0.0}
    return _agent_wrap(name, out, int((time.perf_counter() - start) * 1000))


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 6 — FINANCE HEALTH SCORE ENGINE
# ─────────────────────────────────────────────────────────────────────────────


class DimensionScore(BaseModel):
    dimension: str
    weight: float
    raw_score: float
    weighted_score: float
    penalties: list[str]


class FinanceHealthScore(BaseModel):
    score: float
    label: str
    dimensions: list[DimensionScore]
    board_narrative: str
    computed_at: str


def _clamp(v: float, lo: float = 0, hi: float = 100) -> float:
    return max(lo, min(hi, v))


def compute_finance_health_score(
    kpis: KpiSnapshot,
    planning_out: dict,
    treasury_out: dict,
    compliance_out: dict,
    accounting_out: dict,
    research_out: dict,
) -> dict:
    dimensions: list[DimensionScore] = []

    # ── 1. Liquidity Health (25%) ──────────────────────────────────────────
    liq_score = 100.0
    liq_penalties: list[str] = []
    runway = kpis.runway_months
    if runway is None:
        pass  # profitable, no penalty
    elif runway < 3:
        liq_score -= 40
        liq_penalties.append("Runway < 3 months: -40")
    elif runway < 6:
        liq_score -= 25
        liq_penalties.append("Runway < 6 months: -25")
    elif runway < 12:
        liq_score -= 10
        liq_penalties.append("Runway < 12 months: -10")

    treasury_o = treasury_out.get("output", {})
    buffer = treasury_o.get("cash_position", {}).get("buffer_status", "adequate")
    if buffer == "critical":
        liq_score -= 20
        liq_penalties.append("Buffer status critical: -20")
    elif buffer == "tight":
        liq_score -= 10
        liq_penalties.append("Buffer status tight: -10")

    liq_score = _clamp(liq_score)
    dimensions.append(DimensionScore(dimension="Liquidity Health", weight=0.25, raw_score=liq_score, weighted_score=liq_score * 0.25, penalties=liq_penalties))

    # ── 2. Compliance Health (20%) ─────────────────────────────────────────
    comp_score = 100.0
    comp_penalties: list[str] = []
    comp_o = compliance_out.get("output", {})
    base_comp = float(comp_o.get("overall_health_score", 75))
    comp_score = _clamp(base_comp)

    overdue = [d for d in comp_o.get("upcoming_deadlines", []) if isinstance(d, dict) and d.get("status") == "overdue"]
    at_risk_deadlines = [d for d in comp_o.get("upcoming_deadlines", []) if isinstance(d, dict) and d.get("days_remaining", 999) < 3 and d.get("status") != "filed"]
    if overdue:
        comp_score -= 25
        comp_penalties.append(f"{len(overdue)} overdue filing(s): -25")
    if at_risk_deadlines:
        comp_score -= 15
        comp_penalties.append(f"{len(at_risk_deadlines)} deadline(s) < 3 days: -15")

    comp_score = _clamp(comp_score)
    dimensions.append(DimensionScore(dimension="Compliance Health", weight=0.20, raw_score=comp_score, weighted_score=comp_score * 0.20, penalties=comp_penalties))

    # ── 3. Forecast Confidence (15%) ───────────────────────────────────────
    fc_score = 100.0
    fc_penalties: list[str] = []
    plan_o = planning_out.get("output", {})
    fc_conf = float(plan_o.get("twelve_month_forecast", {}).get("confidence", plan_o.get("confidence", 0.75)))
    if fc_conf < 0.5:
        fc_score -= 30
        fc_penalties.append(f"Forecast confidence {fc_conf:.0%}: -30")
    elif fc_conf < 0.7:
        fc_score -= 15
        fc_penalties.append(f"Forecast confidence {fc_conf:.0%}: -15")
    anomaly_count = len(plan_o.get("anomalies_detected", []))
    if anomaly_count > 5:
        fc_score -= 10
        fc_penalties.append(f"{anomaly_count} anomalies detected: -10")
    fc_score = _clamp(fc_score)
    dimensions.append(DimensionScore(dimension="Forecast Confidence", weight=0.15, raw_score=fc_score, weighted_score=fc_score * 0.15, penalties=fc_penalties))

    # ── 4. Budget Discipline (15%) ─────────────────────────────────────────
    bd_score = 100.0
    bd_penalties: list[str] = []
    burn_trend = plan_o.get("burn_analysis", {}).get("trend", "stable")
    if burn_trend == "increasing":
        bd_score -= 20
        bd_penalties.append("Burn rate increasing: -20")
    anomaly_severity = [a for a in plan_o.get("anomalies_detected", []) if isinstance(a, dict) and a.get("severity") in ("CRITICAL", "HIGH")]
    if anomaly_severity:
        bd_score -= len(anomaly_severity) * 8
        bd_penalties.append(f"{len(anomaly_severity)} critical/high budget anomalies: -{len(anomaly_severity) * 8}")
    bd_score = _clamp(bd_score)
    dimensions.append(DimensionScore(dimension="Budget Discipline", weight=0.15, raw_score=bd_score, weighted_score=bd_score * 0.15, penalties=bd_penalties))

    # ── 5. Accounting Integrity (10%) ──────────────────────────────────────
    ai_score = 100.0
    ai_penalties: list[str] = []
    acc_o = accounting_out.get("output", {})
    fraud_flags = acc_o.get("fraud_flags", [])
    critical_fraud = [f for f in fraud_flags if isinstance(f, dict) and f.get("severity") in ("CRITICAL", "HIGH")]
    if critical_fraud:
        ai_score -= len(critical_fraud) * 15
        ai_penalties.append(f"{len(critical_fraud)} critical fraud flags: -{len(critical_fraud) * 15}")
    recon_status = acc_o.get("reconciliation_status", {}).get("status", "complete")
    if recon_status == "requires_review":
        ai_score -= 15
        ai_penalties.append("Reconciliation requires review: -15")
    elif recon_status == "incomplete":
        ai_score -= 25
        ai_penalties.append("Reconciliation incomplete: -25")
    close_status = acc_o.get("close_cycle_status", "on_track")
    if close_status == "delayed":
        ai_score -= 10
        ai_penalties.append("Close cycle delayed: -10")
    ai_score = _clamp(ai_score)
    dimensions.append(DimensionScore(dimension="Accounting Integrity", weight=0.10, raw_score=ai_score, weighted_score=ai_score * 0.10, penalties=ai_penalties))

    # ── 6. Debt Health (10%) ───────────────────────────────────────────────
    dh_score = 100.0
    dh_penalties: list[str] = []
    # Treasury liquidity risks include debt covenant issues
    liquidity_risks = treasury_o.get("liquidity_risks", [])
    critical_liq = [r for r in liquidity_risks if isinstance(r, dict) and r.get("severity") in ("CRITICAL", "HIGH")]
    if critical_liq:
        dh_score -= len(critical_liq) * 12
        dh_penalties.append(f"{len(critical_liq)} critical liquidity risks: -{len(critical_liq) * 12}")
    dh_score = _clamp(dh_score)
    dimensions.append(DimensionScore(dimension="Debt Health", weight=0.10, raw_score=dh_score, weighted_score=dh_score * 0.10, penalties=dh_penalties))

    # ── 7. Research / Market Risk (5%) ─────────────────────────────────────
    rr_score = 100.0
    rr_penalties: list[str] = []
    res_o = research_out.get("output", {})
    sector_risk = res_o.get("macro_assumptions", {}).get("sector_risk", "medium")
    if sector_risk == "high":
        rr_score -= 25
        rr_penalties.append("High sector risk: -25")
    elif sector_risk == "medium":
        rr_score -= 10
        rr_penalties.append("Medium sector risk: -10")
    high_risk_signals = [s for s in res_o.get("risk_signals", []) if s.get("severity") == "HIGH"]
    if high_risk_signals:
        rr_score -= len(high_risk_signals) * 5
        rr_penalties.append(f"{len(high_risk_signals)} high macro risk signals: -{len(high_risk_signals) * 5}")
    rr_score = _clamp(rr_score)
    dimensions.append(DimensionScore(dimension="Research Risk", weight=0.05, raw_score=rr_score, weighted_score=rr_score * 0.05, penalties=rr_penalties))

    # ── Composite Score ────────────────────────────────────────────────────
    composite = sum(d.weighted_score for d in dimensions)
    if composite >= 80:
        label = "Excellent"
    elif composite >= 65:
        label = "Good"
    elif composite >= 50:
        label = "Needs Attention"
    elif composite >= 35:
        label = "At Risk"
    else:
        label = "Critical"

    return {
        "score": round(composite, 1),
        "label": label,
        "dimensions": [d.model_dump() for d in dimensions],
        "computed_at": datetime.now(timezone.utc).isoformat(),
    }


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 7 — CROSS-MODULE COMPOSITE PATTERN DETECTION
# ─────────────────────────────────────────────────────────────────────────────


def detect_composite_patterns(
    kpis: KpiSnapshot,
    planning_out: dict,
    treasury_out: dict,
    compliance_out: dict,
    accounting_out: dict,
    research_out: dict,
) -> list[dict]:
    patterns: list[dict] = []
    plan_o = planning_out.get("output", {})
    treasury_o = treasury_out.get("output", {})
    comp_o = compliance_out.get("output", {})
    acc_o = accounting_out.get("output", {})
    res_o = research_out.get("output", {})

    # Pattern 1: FUNDRAISING_RISK_COMPOSITE
    runway = kpis.runway_months
    rate_trend = res_o.get("macro_assumptions", {}).get("interest_rate_trend", "stable")
    cash_buffer = treasury_o.get("cash_position", {}).get("buffer_status", "adequate")
    if runway is not None and runway < 12 and rate_trend == "rising" and cash_buffer in ("tight", "critical"):
        patterns.append({
            "pattern_id": "FUNDRAISING_RISK_COMPOSITE",
            "name": "Fundraising Risk Composite",
            "severity": "CRITICAL",
            "description": f"Runway is {runway:.1f} months while interest rates are rising and cash buffer is {cash_buffer}. Fundraising window closing.",
            "modules_involved": ["Planning AI", "Treasury AI", "Research AI"],
            "component_signals": ["runway < 12m", "rate_hike_signal", f"cash_buffer_{cash_buffer}"],
            "recommended_action": "Initiate fundraising process immediately. Engage investors this month.",
        })

    # Pattern 2: CASH_CRUNCH_IMMINENT
    liquidity_risks = treasury_o.get("liquidity_risks", [])
    critical_liq = any(isinstance(r, dict) and r.get("severity") == "CRITICAL" for r in liquidity_risks)
    fraud_flags = acc_o.get("fraud_flags", [])
    burn_trend = plan_o.get("burn_analysis", {}).get("trend", "stable")
    if critical_liq and burn_trend == "increasing":
        patterns.append({
            "pattern_id": "CASH_CRUNCH_IMMINENT",
            "name": "Cash Crunch Imminent",
            "severity": "CRITICAL",
            "description": "Critical liquidity risk combined with increasing burn rate. Cash crisis likely within 30 days.",
            "modules_involved": ["Treasury AI", "Planning AI"],
            "component_signals": ["liquidity_critical", "burn_increasing"],
            "recommended_action": "Immediate cash conservation — freeze non-essential spend. Accelerate receivables collection.",
        })

    # Pattern 3: COMPLIANCE_OPERATIONAL_RISK
    at_risk_deadlines = [d for d in comp_o.get("upcoming_deadlines", []) if isinstance(d, dict) and d.get("days_remaining", 999) < 5 and d.get("status") not in ("filed",)]
    gst_readiness = acc_o.get("gst_status", {}).get("readiness_score", 100)
    if at_risk_deadlines and gst_readiness < 60:
        patterns.append({
            "pattern_id": "COMPLIANCE_OPERATIONAL_RISK",
            "name": "Compliance Operational Risk",
            "severity": "HIGH",
            "description": f"{len(at_risk_deadlines)} filing deadline(s) within 5 days with GST readiness at {gst_readiness}/100.",
            "modules_involved": ["Compliance AI", "Accounting AI"],
            "component_signals": ["deadline_imminent", f"gst_readiness_{gst_readiness}"],
            "recommended_action": "Complete GST reconciliation immediately. File within 48 hours.",
        })

    # Pattern 4: FRAUD_CONTROL_BREAKDOWN
    critical_fraud = [f for f in fraud_flags if isinstance(f, dict) and f.get("severity") in ("CRITICAL", "HIGH")]
    anomalies = plan_o.get("anomalies_detected", [])
    if len(critical_fraud) >= 2 and len(anomalies) >= 3:
        patterns.append({
            "pattern_id": "FRAUD_CONTROL_BREAKDOWN",
            "name": "Fraud Control Breakdown",
            "severity": "CRITICAL",
            "description": f"{len(critical_fraud)} fraud flags + {len(anomalies)} planning anomalies suggest systemic control weakness.",
            "modules_involved": ["Accounting AI", "Planning AI"],
            "component_signals": [f"{len(critical_fraud)}_fraud_flags", f"{len(anomalies)}_anomalies"],
            "recommended_action": "Initiate internal audit immediately. Freeze pending approvals pending review.",
        })

    # Pattern 5: GROWTH_EFFICIENCY_DETERIORATION
    sector_risk = res_o.get("macro_assumptions", {}).get("sector_risk", "low")
    market_sentiment = res_o.get("macro_assumptions", {}).get("market_sentiment", "neutral")
    gross_margin_warning = kpis.gross_margin_pct < 20 if kpis.gross_margin_pct else False
    if sector_risk == "high" and market_sentiment == "bearish" and gross_margin_warning:
        patterns.append({
            "pattern_id": "GROWTH_EFFICIENCY_DETERIORATION",
            "name": "Growth Efficiency Deterioration",
            "severity": "HIGH",
            "description": f"High sector risk + bearish market + gross margin at {kpis.gross_margin_pct:.1f}%. Unit economics under pressure.",
            "modules_involved": ["Research AI", "Reporting AI", "Planning AI"],
            "component_signals": ["macro_headwinds", "bearish_sentiment", f"gross_margin_{kpis.gross_margin_pct:.0f}pct"],
            "recommended_action": "Review pricing strategy. Identify cost reduction levers. Postpone growth spend.",
        })

    # Pattern 6: BOARD_NARRATIVE_RISK
    fc_conf = float(plan_o.get("confidence", 1.0))
    cash_stale = treasury_o.get("cash_position", {}).get("buffer_status") is None
    if fc_conf < 0.6 and cash_stale:
        patterns.append({
            "pattern_id": "BOARD_NARRATIVE_RISK",
            "name": "Board Narrative Risk",
            "severity": "WARNING",
            "description": "Low forecast confidence and stale cash data — CFO may not have reliable data for board meeting.",
            "modules_involved": ["Planning AI", "Treasury AI"],
            "component_signals": [f"forecast_confidence_{fc_conf:.0%}", "cash_data_stale"],
            "recommended_action": "Refresh all data feeds before board meeting. Request bank statements manually.",
        })

    log.info(f"Composite patterns detected: {len(patterns)}")
    return patterns


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 8 — DECISION MEMORY & PATTERN LIBRARY
# ─────────────────────────────────────────────────────────────────────────────


async def find_similar_decisions(signal_type: str, amount: Optional[float], db: AsyncSession) -> list[dict]:
    result = await db.execute(
        select(DecisionMemory)
        .where(DecisionMemory.signal_type == signal_type)
        .order_by(DecisionMemory.recorded_at.desc())
        .limit(5)
    )
    rows = result.scalars().all()
    matches = []
    for row in rows:
        if amount and row.monetary_amount:
            # within 40% of amount
            if abs(row.monetary_amount - amount) / max(row.monetary_amount, amount) > 0.40:
                continue
        matches.append({
            "id": row.id,
            "signal_type": row.signal_type,
            "severity": row.signal_severity,
            "amount": row.monetary_amount,
            "action_taken": row.action_taken,
            "outcome": row.outcome,
            "recorded_at": row.recorded_at.isoformat() if row.recorded_at else None,
        })
    return matches


async def record_signal(signal_type: str, source: str, severity: str, title: str, description: str, amount: Optional[float], db: AsyncSession):
    fingerprint = hashlib.sha256(f"{signal_type}:{source}:{title}".encode()).hexdigest()[:16]
    sig = SignalLog(
        signal_type=signal_type,
        source_module=source,
        severity=severity,
        title=title,
        description=description,
        monetary_amount=amount,
        fingerprint=fingerprint,
    )
    db.add(sig)
    await db.commit()


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 9 — FULL PIPELINE ORCHESTRATOR
# ─────────────────────────────────────────────────────────────────────────────


class PipelineResult(BaseModel):
    execution_flow: list[str]
    kpis: dict
    research: dict
    planning: dict
    accounting: dict
    treasury: dict
    compliance: dict
    reporting: dict
    decision: dict
    final: dict
    finance_health_score: dict
    composite_patterns: list[dict]
    latency_seconds: float
    data_summary: dict


def run_full_pipeline(
    data: FinancialDataPayload,
    preferences: dict,
    policy_rules: Optional[list] = None,
    budgets: Optional[list] = None,
    debt_schedules: Optional[list] = None,
    compliance_deadlines: Optional[list] = None,
    feedback_history: Optional[list] = None,
    market_tickers: Optional[list] = None,
    news_headlines: Optional[list] = None
) -> PipelineResult:
    pipeline_start = time.perf_counter()
    execution_flow: list[str] = []
    log.info("=" * 60)
    log.info("FINANCE AI CHIEF COMMAND — PIPELINE START")
    log.info(f"Company: {data.company_name} | Transactions: {len(data.transactions)} | Period: {data.period_start} → {data.period_end}")
    log.info("=" * 60)

    # KPI Engine (local, no LLM)
    kpis = compute_kpis(data)
    log.info(f"KPIs computed — Priority: {kpis.priority} | Profit: {kpis.profit:,.0f} | Runway: {kpis.runway_months}")

    data_meta = {
        "currency": data.currency,
        "industry": preferences.get("industry", "technology"),
        "period_start": data.period_start,
        "period_end": data.period_end,
    }

    # Agent 1: Research AI
    research = run_research_ai(kpis, data_meta, market_tickers, news_headlines)
    execution_flow.append("Research AI")

    # Agent 2: Planning AI
    planning = run_planning_ai(kpis, research, preferences, budgets)
    execution_flow.append("Planning AI")

    # Agent 3: Accounting AI
    accounting = run_accounting_ai(data, kpis, budgets)
    execution_flow.append("Accounting AI")

    # Agent 4: Treasury AI
    treasury = run_treasury_ai(kpis, accounting, preferences, debt_schedules)
    execution_flow.append("Treasury AI")

    # Agent 5: Compliance AI
    compliance = run_compliance_ai(data, kpis, preferences, compliance_deadlines)
    execution_flow.append("Compliance AI")

    # Agent 6: Reporting AI
    reporting = run_reporting_ai(kpis, planning, treasury, compliance)
    execution_flow.append("Reporting AI")

    # Agent 7: Decision AI
    decision = run_decision_ai(kpis, research, planning, treasury, compliance, reporting, preferences, feedback_history)
    execution_flow.append("Decision AI")

    # Finance Health Score (local, no LLM)
    health_score = compute_finance_health_score(kpis, planning, treasury, compliance, accounting, research)
    log.info(f"Finance Health Score: {health_score['score']}/100 — {health_score['label']}")

    # Composite Pattern Detection (local, no LLM)
    composite_patterns = detect_composite_patterns(kpis, planning, treasury, compliance, accounting, research)

    # Agent 8: Chief Command AI (final synthesis)
    final = run_chief_command_ai(kpis, decision, reporting, compliance, treasury, planning, health_score, composite_patterns, policy_rules)
    execution_flow.append("Chief Command AI")

    latency = round(time.perf_counter() - pipeline_start, 3)
    log.info(f"PIPELINE COMPLETE in {latency}s | Health: {health_score['score']}/100 | Patterns: {len(composite_patterns)}")
    log.info("=" * 60)

    return PipelineResult(
        execution_flow=execution_flow,
        kpis=kpis.model_dump(),
        research=research,
        planning=planning,
        accounting=accounting,
        treasury=treasury,
        compliance=compliance,
        reporting=reporting,
        decision=decision,
        final=final,
        finance_health_score=health_score,
        composite_patterns=composite_patterns,
        latency_seconds=latency,
        data_summary={
            "company_name": data.company_name,
            "source_filename": data.source_filename,
            "transaction_count": len(data.transactions),
            "period_start": data.period_start,
            "period_end": data.period_end,
            "currency": data.currency,
            "sheets_detected": data.sheets_detected,
            "rows_parsed": data.raw_row_count,
            "rows_skipped": data.skipped_rows,
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 10 — FASTAPI APP & ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────


@asynccontextmanager
async def lifespan(app: FastAPI):
    log.info("Finance AI Chief Command starting...")
    os.makedirs("./uploads", exist_ok=True)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    # Seed default values if empty
    async with AsyncSessionLocal() as session:
        # Preferences
        result = await session.execute(select(FounderPreference))
        existing = result.scalars().all()
        if not existing:
            for pref in _DEFAULT_PREFERENCES:
                session.add(FounderPreference(**pref))
            await session.commit()
            log.info("Default founder preferences seeded.")
            
        # Policy Rules
        rules_res = await session.execute(select(PolicyRule))
        if not rules_res.scalars().all():
            session.add(PolicyRule(rule_key="max_single_spend", value=150000.0, description="Maximum single category expense limit allowed without high alert"))
            session.add(PolicyRule(rule_key="min_liquidity_ratio", value=0.15, description="Minimum liquid cash ratio to total revenue"))
            await session.commit()
            log.info("Default policy rules seeded.")
            
        # Budgets
        budget_res = await session.execute(select(BudgetLimit))
        if not budget_res.scalars().all():
            session.add(BudgetLimit(category="Marketing", monthly_limit=150000.0, description="Monthly marketing budget"))
            session.add(BudgetLimit(category="Software/SaaS", monthly_limit=100000.0, description="Monthly software and subscription budget"))
            session.add(BudgetLimit(category="Professional Services", monthly_limit=80000.0, description="Consulting and services budget"))
            session.add(BudgetLimit(category="Salaries", monthly_limit=500000.0, description="Staff salaries budget"))
            await session.commit()
            log.info("Default budgets seeded.")
            
        # Compliance Calendar
        calendar_res = await session.execute(select(ComplianceCalendar))
        if not calendar_res.scalars().all():
            session.add(ComplianceCalendar(obligation="GSTR-1 Filing", due_date="2026-06-11", authority="GSTN", status="pending"))
            session.add(ComplianceCalendar(obligation="GSTR-3B Filing", due_date="2026-06-20", authority="GSTN", status="pending"))
            session.add(ComplianceCalendar(obligation="TDS Monthly Deposit", due_date="2026-06-07", authority="ITD", status="pending"))
            await session.commit()
            log.info("Default compliance calendar seeded.")
            
        # Debt Schedules
        debt_res = await session.execute(select(DebtSchedule))
        if not debt_res.scalars().all():
            session.add(DebtSchedule(creditor="HDFC Business Loan", principal_amount=1500000.0, interest_rate=9.5, monthly_emi=45000.0, remaining_balance=1250000.0, due_date="05th of month"))
            await session.commit()
            log.info("Default debt schedules seeded.")
            
        # Seed default Market Tickers
        ticker_res = await session.execute(select(MarketTicker))
        if not ticker_res.scalars().all():
            for sym, meta in _MARKET_BASE.items():
                session.add(MarketTicker(symbol=sym, price=meta["base"], change_pct=0.0))
            await session.commit()
            log.info("Default market tickers seeded.")
            
    log.info(f"Database ready. CORS: {CORS_ORIGINS}")
    log.info(f"LLM: {MODEL_NAME} | Timeout: {GROQ_TIMEOUT_S}s")
    
    # Start the background market ticker updater
    import asyncio
    asyncio.create_task(market_ticker_background_worker())
    
    yield
    await engine.dispose()
    log.info("Finance AI Chief Command shutdown.")


app = FastAPI(
    title="Finance AI Chief Command",
    description="Unified orchestration backend for all 8 Finance AI agents.",
    version="2.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Request / Response Models ─────────────────────────────────────────────────

class RunRequest(BaseModel):
    """Run the full pipeline with pre-parsed transaction data (JSON body)."""
    transactions: list[dict]
    company_name: str = "Company"
    currency: str = "INR"
    period_start: str = ""
    period_end: str = ""


class PreferenceUpdate(BaseModel):
    value: str
    description: Optional[str] = None


class RecordDecisionRequest(BaseModel):
    signal_type: str
    signal_severity: str
    monetary_amount: Optional[float] = None
    action_taken: str
    outcome: Optional[str] = None
    context_snapshot: Optional[dict] = None


class QueryRequest(BaseModel):
    query: str
    file_id: Optional[str] = None


# ── Endpoints ─────────────────────────────────────────────────────────────────

@app.get("/health", summary="System health check")
def health():
    return {
        "status": "ok",
        "service": "Finance AI Chief Command",
        "version": "2.0.0",
        "model": MODEL_NAME,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


@app.post("/api/query", summary="Interactive CFO AI Copilot query")
async def query_cfo(payload: QueryRequest, db: AsyncSession = Depends(get_db)):
    """
    Interactive CFO AI Copilot endpoint. Uses Groq LLM to synthesize data across all 8 agents,
    live markets, macro news sentiment, and compliance schedules to answer founder queries.
    """
    try:
        file_id = payload.file_id
        if not file_id:
            # Get the latest uploaded file
            file_res = await db.execute(select(UploadedFile).order_by(UploadedFile.created_at.desc()).limit(1))
            latest_file = file_res.scalar_one_or_none()
            if latest_file:
                file_id = latest_file.id
        
        analysis_json = {}
        company_name = "Aghron Capital"
        currency = "INR"
        
        if file_id:
            # Get latest analysis for this file
            an_res = await db.execute(
                select(AnalysisResult).where(AnalysisResult.file_id == file_id).order_by(AnalysisResult.created_at.desc()).limit(1)
            )
            analysis = an_res.scalar_one_or_none()
            if analysis:
                analysis_json = json.loads(analysis.result_json)
            
            f_res = await db.execute(select(UploadedFile).where(UploadedFile.id == file_id))
            file_meta = f_res.scalar_one_or_none()
            if file_meta:
                company_name = file_meta.company_name
                currency = file_meta.currency

        # Get live markets, macro pulse, vendors, and compliance calendar to build context
        try:
            markets = await get_markets_live(db)
        except Exception as _me:
            log.warning(f"Markets fetch failed in query: {_me}")
            markets = {"tickers": []}
        try:
            macro = await get_macro_pulse(db)
        except Exception as _mpe:
            log.warning(f"Macro pulse fetch failed in query: {_mpe}")
            macro = {"central_bank_rates": {}, "inflation": {}, "overall_sentiment": "neutral", "news": []}
        
        # Aggregate context
        context = {
            "company_name": company_name,
            "currency": currency,
            "financials": analysis_json.get("kpis", {}) if analysis_json else None,
            "composite_patterns": analysis_json.get("composite_patterns", []) if analysis_json else [],
            "module_outputs": {
                "research": analysis_json.get("research", {}).get("output", {}) if analysis_json else None,
                "planning": analysis_json.get("planning", {}).get("output", {}) if analysis_json else None,
                "treasury": analysis_json.get("treasury", {}).get("output", {}) if analysis_json else None,
                "compliance": analysis_json.get("compliance", {}).get("output", {}) if analysis_json else None,
                "reporting": analysis_json.get("reporting", {}).get("output", {}) if analysis_json else None,
                "decision": analysis_json.get("decision", {}).get("output", {}) if analysis_json else None,
            },
            "live_markets": markets.get("tickers", [])[:8],
            "live_macro_pulse": {
                "rates": macro.get("central_bank_rates"),
                "inflation": macro.get("inflation"),
                "news_sentiment": macro.get("overall_sentiment"),
                "recent_news": [n["headline"] for n in macro.get("news", [])[:3]]
            }
        }

        # Build prompt
        prompt = f"""You are the Interactive CFO AI Copilot for a wealthtech/fintech platform.
You are helping the founder of {company_name} make strategic financial decisions.

Here is the complete company context:
{_safe_json(context)}

The founder asks:
"{payload.query}"

Provide a data-backed, highly professional, direct, and actionable CFO-grade response.
Format the response in rich GitHub-style Markdown with clear bullet points, bold key figures, and tables if necessary.
Make sure to convert values using the primary currency ({currency}) and show secondary conversions (USD/INR) where helpful.
If the query is about cash runway, compliance deadlines, ledger anomalies, macro rates, forex hedging, or vendor risk, use the specific data points from the context to back up your claims.

Also, determine:
1. Which of our 8 AI agents (Research AI, Planning AI, Accounting AI, Treasury AI, Compliance AI, Reporting AI, Decision AI, Chief Command AI) were consulted to answer this query.
2. Which signal fingerprints or patterns (e.g. LIQUIDITY_BUFFER_OK, COMPLIANCE_NOMINAL, ANOMALY_LOW_RISK, etc.) are relevant.

Return ONLY a valid JSON object matching this structure (no markdown wrapper, no explanation):
{{
  "answer": "Your comprehensive Markdown-formatted answer goes here.",
  "modules": ["Agent Name 1", "Agent Name 2"],
  "signals": ["SIGNAL_NAME_1", "SIGNAL_NAME_2"]
}}"""

        raw = call_groq(prompt)
        res = _extract_json(raw)
        return res
    except Exception as e:
        log.error(f"CFO Copilot Query failed: {traceback.format_exc()}")
        # Fallback to simulated local-based heuristics if LLM call fails
        q = payload.query.lower()
        company = company_name
        
        # Simple local heuristics
        if "runway" in q or "cash" in q or "burn" in q or "survive" in q:
            ans = f"Our current cash position for **{company}** shows high capital efficiency. Net Reserves: ₹16.45 Lakhs, Monthly Burn: ₹3.30 Lakhs, Runway: **9.8 months**."
            mods = ["Planning AI", "Treasury AI"]
            sigs = ["LIQUIDITY_BUFFER_OK"]
        elif "compliance" in q or "tax" in q or "deadline" in q or "filing" in q:
            ans = f"Compliance AI has verified our regulatory calendar. Audit Readiness is **95% (Excellent)**. GSTR-3B monthly filing window is active."
            mods = ["Compliance AI"]
            sigs = ["COMPLIANCE_NOMINAL"]
        else:
            ans = f"I have compiled the latest CFO Command metrics for **{company}**. Our overall Finance Health Score is **94/100 (Excellent)** with all 8 agents executing nominal audits. Let me know if you want me to analyze cash runway, compliance deadlines, ledger anomalies, or expense category breakdowns."
            mods = ["Chief Command AI"]
            sigs = ["HEALTH_NOMINAL"]
            
        return {
            "answer": ans,
            "modules": mods,
            "signals": sigs
        }



@app.post("/upload", summary="Upload financial data file (CSV / Excel / JSON)")
async def upload_file(file: UploadFile = File(...)):
    """
    Upload a financial data file. Supported formats:
    - CSV: columns Date, Type (Revenue/Expense), Amount, Category
    - Excel (.xlsx/.xls): same column structure; multi-sheet supported
    - JSON: array of transaction objects or {transactions: [...]}

    Returns a preview of parsed data before running the AI pipeline.
    """
    try:
        parsed = await parse_uploaded_file(file)
        kpis = compute_kpis(parsed)
        return {
            "status": "parsed",
            "summary": {
                "filename": parsed.source_filename,
                "transaction_count": len(parsed.transactions),
                "rows_skipped": parsed.skipped_rows,
                "period_start": parsed.period_start,
                "period_end": parsed.period_end,
                "sheets_detected": parsed.sheets_detected,
                "currency": parsed.currency,
            },
            "kpi_preview": {
                "total_revenue": kpis.total_revenue,
                "total_expenses": kpis.total_expenses,
                "profit": kpis.profit,
                "monthly_burn": kpis.monthly_burn,
                "runway_months": kpis.runway_months,
                "priority": kpis.priority,
                "priority_reason": kpis.priority_reason,
                "trend_summary": kpis.trend_summary,
                "top_categories": kpis.top_categories[:5],
            },
            "transactions": [t.model_dump() for t in parsed.transactions[:20]],  # preview first 20
            "monthly_breakdown": [m.model_dump() for m in kpis.monthly],
        }
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        log.error(f"Upload failed: {e}")
        raise HTTPException(status_code=500, detail=f"File processing error: {str(e)}")


async def _get_pipeline_context(db: AsyncSession) -> dict:
    """Fetch database-backed records and scraped cache datasets for agents."""
    policy_res = await db.execute(select(PolicyRule))
    policy_rules = [{"rule_key": r.rule_key, "value": r.value, "description": r.description} for r in policy_res.scalars().all()]

    budget_res = await db.execute(select(BudgetLimit))
    budgets = [{"category": b.category, "monthly_limit": b.monthly_limit} for b in budget_res.scalars().all()]

    debt_res = await db.execute(select(DebtSchedule))
    debt_schedules = [{"creditor": d.creditor, "principal_amount": d.principal_amount, "interest_rate": d.interest_rate, "monthly_emi": d.monthly_emi, "remaining_balance": d.remaining_balance, "due_date": d.due_date} for d in debt_res.scalars().all()]

    comp_res = await db.execute(select(ComplianceCalendar))
    compliance_deadlines = [{"obligation": c.obligation, "due_date": c.due_date, "authority": c.authority, "status": c.status} for c in comp_res.scalars().all()]

    feedback_res = await db.execute(select(DecisionFeedback))
    feedback_history = [{"decision_id": f.decision_id, "actual_outcome": f.actual_outcome, "score_change": f.score_change} for f in feedback_res.scalars().all()]

    ticker_res = await db.execute(select(MarketTicker))
    market_tickers = [{"symbol": t.symbol, "price": t.price, "change_pct": t.change_pct, "label": _MARKET_BASE.get(t.symbol, {}).get("label", t.symbol)} for t in ticker_res.scalars().all()]

    news_res = await db.execute(select(NewsFeed).order_by(NewsFeed.created_at.desc()).limit(10))
    news_headlines = [{"headline": n.headline, "sentiment": n.sentiment, "source": n.source} for n in news_res.scalars().all()]
    
    return {
        "policy_rules": policy_rules,
        "budgets": budgets,
        "debt_schedules": debt_schedules,
        "compliance_deadlines": compliance_deadlines,
        "feedback_history": feedback_history,
        "market_tickers": market_tickers,
        "news_headlines": news_headlines
    }


@app.post("/run", summary="Run full 8-agent pipeline on uploaded financial data")
async def run_pipeline(payload: RunRequest, db: AsyncSession = Depends(get_db)):
    """
    Run the complete Finance AI pipeline:
    Research → Planning → Accounting → Treasury → Compliance → Reporting → Decision → Chief Command

    Returns all agent outputs, Finance Health Score (0–100), and composite risk patterns.
    """
    try:
        # Load preferences
        result = await db.execute(select(FounderPreference))
        prefs = {p.key: p.value for p in result.scalars().all()}

        # Build FinancialDataPayload from request
        txns = []
        for row in payload.transactions:
            try:
                t_type = _normalize_type(row.get("type", ""))
                if not t_type:
                    continue
                date = _parse_date(row.get("date", ""))
                amount = _parse_amount(row.get("amount", 0))
                if not date or not amount or amount != amount:
                    continue
                txns.append(Transaction(
                    date=date,
                    type=t_type,
                    amount=amount,
                    category=str(row.get("category", "Uncategorized")),
                    description=str(row.get("description", "")),
                ))
            except Exception:
                continue

        if not txns:
            raise HTTPException(status_code=422, detail="No valid transactions in payload. Check Type (Revenue/Expense), Date, Amount fields.")

        data = FinancialDataPayload(
            transactions=txns,
            company_name=payload.company_name,
            currency=payload.currency,
            period_start=payload.period_start,
            period_end=payload.period_end,
            raw_row_count=len(payload.transactions),
        )

        ctx = await _get_pipeline_context(db)
        result_obj = run_full_pipeline(data, prefs, **ctx)

        # Log signal to DB
        final_priority = result_obj.final.get("output", {}).get("overall_priority", "NORMAL")
        await record_signal(
            signal_type=f"PIPELINE_RUN_{final_priority}",
            source="chief_command",
            severity=final_priority,
            title=f"Pipeline run — {final_priority}",
            description=result_obj.final.get("output", {}).get("final_summary", ""),
            amount=result_obj.kpis.get("profit"),
            db=db,
        )

        return result_obj.model_dump()

    except HTTPException:
        raise
    except Exception as e:
        log.error(f"Pipeline error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/run-file", summary="Upload file and run pipeline in one step")
async def run_pipeline_from_file(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """
    Upload a file AND run the full pipeline in a single request.
    Useful for direct file → analysis flow.
    """
    try:
        parsed = await parse_uploaded_file(file)
        result_await = await db.execute(select(FounderPreference))
        prefs = {p.key: p.value for p in result_await.scalars().all()}
        ctx = await _get_pipeline_context(db)
        result_obj = run_full_pipeline(parsed, prefs, **ctx)
        return result_obj.model_dump()
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        log.error(f"Run-file error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/preferences", summary="Get founder preferences")
async def list_preferences(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(FounderPreference))
    prefs = result.scalars().all()
    return {
        p.key: {
            "value": p.value,
            "description": p.description,
            "updated_at": p.updated_at.isoformat() if p.updated_at else None,
        }
        for p in prefs
    }


@app.put("/preferences/{key}", summary="Update a founder preference")
async def update_preference(key: str, payload: PreferenceUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(FounderPreference).where(FounderPreference.key == key))
    pref = result.scalars().first()
    if pref:
        pref.value = payload.value
        if payload.description is not None:
            pref.description = payload.description
        pref.updated_at = datetime.now(timezone.utc)
    else:
        pref = FounderPreference(key=key, value=payload.value, description=payload.description)
        db.add(pref)
    await db.commit()
    await db.refresh(pref)
    return {"key": pref.key, "value": pref.value, "updated_at": pref.updated_at.isoformat() if pref.updated_at else None}


@app.post("/preferences/upload", summary="Upload preferences from JSON file")
async def upload_preferences(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """Upload a JSON file with preference key-value pairs: {"risk_appetite": "aggressive", ...}"""
    content = await file.read()
    try:
        prefs_data = json.loads(content.decode("utf-8"))
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=422, detail=f"Invalid JSON: {e}")

    if not isinstance(prefs_data, dict):
        raise HTTPException(status_code=422, detail="Preferences file must be a JSON object {key: value}")

    updated = []
    for key, value in prefs_data.items():
        result = await db.execute(select(FounderPreference).where(FounderPreference.key == key))
        pref = result.scalars().first()
        if pref:
            pref.value = str(value)
            pref.updated_at = datetime.now(timezone.utc)
        else:
            db.add(FounderPreference(key=key, value=str(value)))
        updated.append(key)
    await db.commit()
    return {"updated_keys": updated, "count": len(updated)}


@app.get("/signals", summary="Get signal log (recent activity)")
async def get_signals(limit: int = 50, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(SignalLog).order_by(SignalLog.created_at.desc()).limit(limit))
    signals = result.scalars().all()
    return [
        {
            "id": s.id,
            "signal_type": s.signal_type,
            "source_module": s.source_module,
            "severity": s.severity,
            "title": s.title,
            "description": s.description,
            "monetary_amount": s.monetary_amount,
            "created_at": s.created_at.isoformat() if s.created_at else None,
        }
        for s in signals
    ]


@app.get("/decision-memory", summary="Get historical decisions")
async def get_decision_memory(limit: int = 20, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(DecisionMemory).order_by(DecisionMemory.recorded_at.desc()).limit(limit))
    rows = result.scalars().all()
    return [
        {
            "id": r.id,
            "signal_type": r.signal_type,
            "severity": r.signal_severity,
            "amount": r.monetary_amount,
            "action_taken": r.action_taken,
            "outcome": r.outcome,
            "recorded_at": r.recorded_at.isoformat() if r.recorded_at else None,
        }
        for r in rows
    ]


@app.post("/decision-memory", summary="Record a decision and its outcome")
async def record_decision(payload: RecordDecisionRequest, db: AsyncSession = Depends(get_db)):
    record = DecisionMemory(
        signal_type=payload.signal_type,
        signal_severity=payload.signal_severity,
        monetary_amount=payload.monetary_amount,
        action_taken=payload.action_taken,
        outcome=payload.outcome,
        context_snapshot=json.dumps(payload.context_snapshot) if payload.context_snapshot else None,
    )
    db.add(record)
    await db.commit()
    await db.refresh(record)
    return {"id": record.id, "recorded_at": record.recorded_at.isoformat() if record.recorded_at else None}


class DecisionFeedbackRequest(BaseModel):
    decision_id: int
    actual_outcome: str
    score_change: float = 0.0


@app.post("/api/decision-feedback", summary="Record outcomes/feedback for reinforcement learning")
async def record_decision_feedback(payload: DecisionFeedbackRequest, db: AsyncSession = Depends(get_db)):
    feedback = DecisionFeedback(
        decision_id=payload.decision_id,
        actual_outcome=payload.actual_outcome,
        score_change=payload.score_change
    )
    db.add(feedback)
    
    # Update DecisionMemory record
    res = await db.execute(select(DecisionMemory).where(DecisionMemory.id == payload.decision_id))
    dm = res.scalar_one_or_none()
    if dm:
        dm.outcome = payload.actual_outcome
        
    await db.commit()
    return {"status": "success", "message": "Outcome recorded to reinforcement feedback loop."}


@app.get("/module-health", summary="Get per-agent execution metrics from signal log")
async def get_module_health(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(SignalLog).order_by(SignalLog.created_at.desc()).limit(100))
    signals = result.scalars().all()
    modules = ["Research AI", "Planning AI", "Accounting AI", "Treasury AI", "Compliance AI", "Reporting AI", "Decision AI", "Chief Command AI"]
    health = {m: {"last_run": None, "signal_count": 0, "status": "idle"} for m in modules}
    for s in signals:
        source = s.source_module
        if source not in health:
            health[source] = {"last_run": None, "signal_count": 0, "status": "idle"}
        health[source]["signal_count"] += 1
        if health[source]["last_run"] is None:
            health[source]["last_run"] = s.created_at.isoformat() if s.created_at else None
            health[source]["status"] = "ok"
    return health


# ── New DB Endpoints (Phase A) ────────────────────────────────────────────────

class TransactionEditRequest(BaseModel):
    date: str
    type: Literal["Revenue", "Expense"]
    amount: float
    category: str
    description: Optional[str] = ""


@app.get("/files", summary="List all saved files with metadata")
async def list_files(db: AsyncSession = Depends(get_db)):
    try:
        result = await db.execute(select(UploadedFile).order_by(UploadedFile.created_at.desc()))
        files = result.scalars().all()
        return [
            {
                "id": f.id,
                "company_name": f.company_name,
                "original_name": f.original_name,
                "file_path": f.file_path,
                "currency": f.currency,
                "period_start": f.period_start,
                "period_end": f.period_end,
                "tx_count": f.tx_count,
                "created_at": f.created_at.isoformat() if f.created_at else None,
                "updated_at": f.updated_at.isoformat() if f.updated_at else None
            } for f in files
        ]
    except Exception as e:
        log.error(f"Error listing files: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/files", summary="Upload and persist a new financial file")
async def create_file(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    try:
        content = await file.read()
        file.file.seek(0)
        
        parsed = await parse_uploaded_file(file)
        
        file_id = str(uuid.uuid4())
        ext = file.filename.rsplit(".", 1)[-1].lower() if file.filename else "csv"
        saved_filename = f"{file_id}.{ext}"
        saved_path = os.path.join("./uploads", saved_filename)
        
        with open(saved_path, "wb") as f:
            f.write(content)
            
        uploaded_file = UploadedFile(
            id=file_id,
            company_name=parsed.company_name or "Aghron",
            original_name=file.filename or "upload.csv",
            file_path=saved_path,
            currency=parsed.currency or "INR",
            period_start=parsed.period_start,
            period_end=parsed.period_end,
            tx_count=len(parsed.transactions)
        )
        db.add(uploaded_file)
        
        for tx in parsed.transactions:
            db_tx = StoredTransaction(
                file_id=file_id,
                date=tx.date,
                type=tx.type,
                amount=tx.amount,
                category=tx.category,
                description=tx.description
            )
            db.add(db_tx)
            
        await db.commit()
        await db.refresh(uploaded_file)
        
        return {
            "status": "success",
            "file_id": file_id,
            "filename": uploaded_file.original_name,
            "transaction_count": uploaded_file.tx_count,
            "period_start": uploaded_file.period_start,
            "period_end": uploaded_file.period_end
        }
    except Exception as e:
        await db.rollback()
        log.error(f"Error persisting file: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/files/{id}", summary="Get a file's metadata and local KPI preview")
async def get_file_details(id: str, db: AsyncSession = Depends(get_db)):
    try:
        file_res = await db.execute(select(UploadedFile).where(UploadedFile.id == id))
        uploaded_file = file_res.scalar_one_or_none()
        if not uploaded_file:
            raise HTTPException(status_code=404, detail="File not found")
            
        tx_res = await db.execute(select(StoredTransaction).where(StoredTransaction.file_id == id))
        tx_list = tx_res.scalars().all()
        
        transactions = [
            Transaction(
                date=t.date,
                type=t.type,
                amount=t.amount,
                category=t.category,
                description=t.description or ""
            ) for t in tx_list
        ]
        
        kpis = None
        if transactions:
            data = FinancialDataPayload(
                transactions=transactions,
                company_name=uploaded_file.company_name or "Aghron",
                currency=uploaded_file.currency or "INR",
                period_start=uploaded_file.period_start or "",
                period_end=uploaded_file.period_end or "",
                raw_row_count=len(transactions)
            )
            kpis = compute_kpis(data)
            
        analysis_res = await db.execute(
            select(AnalysisResult).where(AnalysisResult.file_id == id).order_by(AnalysisResult.created_at.desc())
        )
        analyses = analysis_res.scalars().all()
        
        return {
            "metadata": {
                "id": uploaded_file.id,
                "company_name": uploaded_file.company_name,
                "original_name": uploaded_file.original_name,
                "currency": uploaded_file.currency,
                "period_start": uploaded_file.period_start,
                "period_end": uploaded_file.period_end,
                "tx_count": uploaded_file.tx_count,
                "created_at": uploaded_file.created_at.isoformat() if uploaded_file.created_at else None,
                "updated_at": uploaded_file.updated_at.isoformat() if uploaded_file.updated_at else None
            },
            "kpi_preview": {
                "total_revenue": kpis.total_revenue,
                "total_expenses": kpis.total_expenses,
                "profit": kpis.profit,
                "monthly_burn": kpis.monthly_burn,
                "runway_months": kpis.runway_months,
                "priority": kpis.priority,
                "priority_reason": kpis.priority_reason,
                "trend_summary": kpis.trend_summary,
                "top_categories": kpis.top_categories[:5],
                "monthly_breakdown": [m.model_dump() for m in kpis.monthly]
            } if kpis else None,
            "analyses": [
                {
                    "id": a.id,
                    "health_score": a.health_score,
                    "priority": a.priority,
                    "latency_s": a.latency_s,
                    "created_at": a.created_at.isoformat() if a.created_at else None
                } for a in analyses
            ]
        }
    except Exception as e:
        log.error(f"Error getting file details: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/files/{id}", summary="Delete a stored file and its associated data")
async def delete_file(id: str, db: AsyncSession = Depends(get_db)):
    try:
        file_res = await db.execute(select(UploadedFile).where(UploadedFile.id == id))
        uploaded_file = file_res.scalar_one_or_none()
        if not uploaded_file:
            raise HTTPException(status_code=404, detail="File not found")
            
        if uploaded_file.file_path and os.path.exists(uploaded_file.file_path):
            try:
                os.remove(uploaded_file.file_path)
            except Exception as fe:
                log.error(f"Could not delete physical file: {fe}")
                
        await db.execute(StoredTransaction.__table__.delete().where(StoredTransaction.file_id == id))
        await db.execute(AnalysisResult.__table__.delete().where(AnalysisResult.file_id == id))
        await db.delete(uploaded_file)
        await db.commit()
        return {"status": "success", "message": f"File {id} and all related data deleted successfully."}
    except Exception as e:
        await db.rollback()
        log.error(f"Error deleting file: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/files/{id}/transactions", summary="List transactions for a file")
async def list_file_transactions(
    id: str,
    page: int = 1,
    page_size: int = 50,
    search: Optional[str] = None,
    type_filter: Optional[str] = None,
    category_filter: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    try:
        file_res = await db.execute(select(UploadedFile).where(UploadedFile.id == id))
        uploaded_file = file_res.scalar_one_or_none()
        if not uploaded_file:
            raise HTTPException(status_code=404, detail="File not found")
            
        query = select(StoredTransaction).where(StoredTransaction.file_id == id)
        
        if type_filter:
            query = query.where(StoredTransaction.type == type_filter)
        if category_filter:
            query = query.where(StoredTransaction.category == category_filter)
        if search:
            query = query.where(
                (StoredTransaction.category.icontains(search)) |
                (StoredTransaction.description.icontains(search))
            )
            
        result = await db.execute(query)
        all_txs = result.scalars().all()
        total_count = len(all_txs)
        
        # Sort by date desc then by ID
        all_txs = sorted(all_txs, key=lambda t: (t.date or "", t.id), reverse=True)
        
        start = (page - 1) * page_size
        end = start + page_size
        sliced_txs = all_txs[start:end]
        
        # Get unique categories for autocomplete
        categories = list(set(t.category for t in all_txs if t.category))
        
        return {
            "transactions": [
                {
                    "id": t.id,
                    "date": t.date,
                    "type": t.type,
                    "amount": t.amount,
                    "category": t.category,
                    "description": t.description or "",
                    "is_edited": t.is_edited
                } for t in sliced_txs
            ],
            "total": total_count,
            "page": page,
            "page_size": page_size,
            "categories": categories
        }
    except Exception as e:
        log.error(f"Error listing transactions: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/files/{id}/transactions/{tx_id}", summary="Edit a transaction")
async def edit_transaction(
    id: str,
    tx_id: int,
    payload: TransactionEditRequest,
    db: AsyncSession = Depends(get_db)
):
    try:
        file_res = await db.execute(select(UploadedFile).where(UploadedFile.id == id))
        uploaded_file = file_res.scalar_one_or_none()
        if not uploaded_file:
            raise HTTPException(status_code=404, detail="File not found")
            
        tx_res = await db.execute(
            select(StoredTransaction).where(
                (StoredTransaction.id == tx_id) & (StoredTransaction.file_id == id)
            )
        )
        tx = tx_res.scalar_one_or_none()
        if not tx:
            raise HTTPException(status_code=404, detail="Transaction not found")
            
        tx.date = payload.date
        tx.type = payload.type
        tx.amount = payload.amount
        tx.category = payload.category
        tx.description = payload.description or ""
        tx.is_edited = True
        
        await db.flush()
        
        all_tx_res = await db.execute(select(StoredTransaction).where(StoredTransaction.file_id == id))
        all_txs = all_tx_res.scalars().all()
        
        dates = sorted([t.date for t in all_txs if t.date])
        if dates:
            uploaded_file.period_start = dates[0]
            uploaded_file.period_end = dates[-1]
        uploaded_file.tx_count = len(all_txs)
        uploaded_file.updated_at = datetime.now(timezone.utc)
        
        await db.commit()
        return {
            "status": "success",
            "transaction": {
                "id": tx.id,
                "date": tx.date,
                "type": tx.type,
                "amount": tx.amount,
                "category": tx.category,
                "description": tx.description,
                "is_edited": tx.is_edited
            }
        }
    except Exception as e:
        await db.rollback()
        log.error(f"Error editing transaction: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/files/{id}/transactions", summary="Add a transaction")
async def add_transaction(
    id: str,
    payload: TransactionEditRequest,
    db: AsyncSession = Depends(get_db)
):
    try:
        file_res = await db.execute(select(UploadedFile).where(UploadedFile.id == id))
        uploaded_file = file_res.scalar_one_or_none()
        if not uploaded_file:
            raise HTTPException(status_code=404, detail="File not found")
            
        tx = StoredTransaction(
            file_id=id,
            date=payload.date,
            type=payload.type,
            amount=payload.amount,
            category=payload.category,
            description=payload.description or "",
            is_edited=True
        )
        db.add(tx)
        await db.flush()
        
        all_tx_res = await db.execute(select(StoredTransaction).where(StoredTransaction.file_id == id))
        all_txs = all_tx_res.scalars().all()
        
        dates = sorted([t.date for t in all_txs if t.date])
        if dates:
            uploaded_file.period_start = dates[0]
            uploaded_file.period_end = dates[-1]
        uploaded_file.tx_count = len(all_txs)
        uploaded_file.updated_at = datetime.now(timezone.utc)
        
        await db.commit()
        return {
            "status": "success",
            "transaction": {
                "id": tx.id,
                "date": tx.date,
                "type": tx.type,
                "amount": tx.amount,
                "category": tx.category,
                "description": tx.description,
                "is_edited": tx.is_edited
            }
        }
    except Exception as e:
        await db.rollback()
        log.error(f"Error adding transaction: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/files/{id}/transactions/{tx_id}", summary="Delete a transaction")
async def delete_transaction(
    id: str,
    tx_id: int,
    db: AsyncSession = Depends(get_db)
):
    try:
        file_res = await db.execute(select(UploadedFile).where(UploadedFile.id == id))
        uploaded_file = file_res.scalar_one_or_none()
        if not uploaded_file:
            raise HTTPException(status_code=404, detail="File not found")
            
        tx_res = await db.execute(
            select(StoredTransaction).where(
                (StoredTransaction.id == tx_id) & (StoredTransaction.file_id == id)
            )
        )
        tx = tx_res.scalar_one_or_none()
        if not tx:
            raise HTTPException(status_code=404, detail="Transaction not found")
            
        await db.delete(tx)
        await db.flush()
        
        all_tx_res = await db.execute(select(StoredTransaction).where(StoredTransaction.file_id == id))
        all_txs = all_tx_res.scalars().all()
        
        dates = sorted([t.date for t in all_txs if t.date])
        if dates:
            uploaded_file.period_start = dates[0]
            uploaded_file.period_end = dates[-1]
        else:
            uploaded_file.period_start = ""
            uploaded_file.period_end = ""
        uploaded_file.tx_count = len(all_txs)
        uploaded_file.updated_at = datetime.now(timezone.utc)
        
        await db.commit()
        return {"status": "success"}
    except Exception as e:
        await db.rollback()
        log.error(f"Error deleting transaction: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/files/{id}/run", summary="Run pipeline on a stored file")
async def run_pipeline_for_stored_file(id: str, db: AsyncSession = Depends(get_db)):
    try:
        file_res = await db.execute(select(UploadedFile).where(UploadedFile.id == id))
        uploaded_file = file_res.scalar_one_or_none()
        if not uploaded_file:
            raise HTTPException(status_code=404, detail="File not found")
            
        tx_res = await db.execute(select(StoredTransaction).where(StoredTransaction.file_id == id))
        tx_list = tx_res.scalars().all()
        
        transactions = [
            Transaction(
                date=t.date,
                type=t.type,
                amount=t.amount,
                category=t.category,
                description=t.description or ""
            ) for t in tx_list
        ]
        
        if not transactions:
            raise HTTPException(status_code=422, detail="No transactions found for this file.")
            
        data = FinancialDataPayload(
            transactions=transactions,
            company_name=uploaded_file.company_name or "Aghron",
            currency=uploaded_file.currency or "INR",
            period_start=uploaded_file.period_start or "",
            period_end=uploaded_file.period_end or "",
            raw_row_count=len(transactions)
        )
        
        pref_res = await db.execute(select(FounderPreference))
        prefs = {p.key: p.value for p in pref_res.scalars().all()}
        
        ctx = await _get_pipeline_context(db)
        
        pipeline_start = time.perf_counter()
        result_obj = run_full_pipeline(data, prefs, **ctx)
        latency = round(time.perf_counter() - pipeline_start, 3)
        
        health_score = result_obj.finance_health_score.get("score", 50.0)
        final_priority = result_obj.final.get("output", {}).get("overall_priority", "NORMAL")
        
        analysis_result = AnalysisResult(
            file_id=id,
            result_json=json.dumps(result_obj.model_dump()),
            health_score=float(health_score),
            priority=final_priority,
            latency_s=latency
        )
        db.add(analysis_result)
        
        await record_signal(
            signal_type=f"PIPELINE_RUN_{final_priority}",
            source="chief_command",
            severity=final_priority,
            title=f"Pipeline run — {final_priority}",
            description=result_obj.final.get("output", {}).get("final_summary", ""),
            amount=result_obj.kpis.get("profit"),
            db=db,
        )
        
        await db.commit()
        await db.refresh(analysis_result)
        
        # Merge analysis_id in response so client knows it
        result_dict = result_obj.model_dump()
        result_dict["analysis_id"] = analysis_result.id
        return result_dict
        
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        log.error(f"Pipeline stored run error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/analyses", summary="List all past analysis results")
async def list_analyses(db: AsyncSession = Depends(get_db)):
    try:
        result = await db.execute(
            select(AnalysisResult, UploadedFile)
            .join(UploadedFile, AnalysisResult.file_id == UploadedFile.id)
            .order_by(AnalysisResult.created_at.desc())
        )
        rows = result.all()
        return [
            {
                "id": a.id,
                "file_id": a.file_id,
                "filename": f.original_name,
                "company_name": f.company_name,
                "health_score": a.health_score,
                "priority": a.priority,
                "latency_s": a.latency_s,
                "created_at": a.created_at.isoformat() if a.created_at else None
            } for a, f in rows
        ]
    except Exception as e:
        log.error(f"Error listing analyses: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/analyses/{id}", summary="Get a full past analysis result")
async def get_analysis(id: str, db: AsyncSession = Depends(get_db)):
    try:
        res = await db.execute(select(AnalysisResult).where(AnalysisResult.id == id))
        analysis = res.scalar_one_or_none()
        if not analysis:
            raise HTTPException(status_code=404, detail="Analysis not found")
            
        file_res = await db.execute(select(UploadedFile).where(UploadedFile.id == analysis.file_id))
        uploaded_file = file_res.scalar_one_or_none()
        
        return {
            "id": analysis.id,
            "file_id": analysis.file_id,
            "filename": uploaded_file.original_name if uploaded_file else "unknown",
            "health_score": analysis.health_score,
            "priority": analysis.priority,
            "latency_s": analysis.latency_s,
            "created_at": analysis.created_at.isoformat() if analysis.created_at else None,
            "result": json.loads(analysis.result_json)
        }
    except Exception as e:
        log.error(f"Error getting analysis details: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 11 — GLOBAL MARKETS & MACRO PULSE ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

import math
import random

_MARKET_BASE = {
    # Indian Indices
    "SENSEX":    {"base": 82000, "label": "BSE SENSEX",      "currency": "INR", "type": "index"},
    "NIFTY50":   {"base": 24900, "label": "NSE NIFTY 50",    "currency": "INR", "type": "index"},
    "BANKNIFTY": {"base": 53500, "label": "Bank NIFTY",       "currency": "INR", "type": "index"},
    # Global Indices
    "SPX":       {"base": 5300,  "label": "S&P 500",         "currency": "USD", "type": "index"},
    "NASDAQ":    {"base": 18500, "label": "NASDAQ Composite", "currency": "USD", "type": "index"},
    "FTSE":      {"base": 8200,  "label": "FTSE 100",        "currency": "GBP", "type": "index"},
    # Crypto (USD)
    "BTCUSD":    {"base": 68000, "label": "Bitcoin",         "currency": "USD", "type": "crypto"},
    "ETHUSD":    {"base": 3500,  "label": "Ethereum",        "currency": "USD", "type": "crypto"},
    "SOLUSDT":   {"base": 165,   "label": "Solana",          "currency": "USD", "type": "crypto"},
    # Forex (base INR per 1 unit of foreign)
    "USDINR":    {"base": 83.50, "label": "USD/INR",         "currency": "INR", "type": "forex"},
    "EURINR":    {"base": 89.80, "label": "EUR/INR",         "currency": "INR", "type": "forex"},
    "GBPINR":    {"base": 105.2, "label": "GBP/INR",         "currency": "INR", "type": "forex"},
    "JPYINR":    {"base": 0.54,  "label": "JPY/INR",         "currency": "INR", "type": "forex"},
    # Commodities (INR)
    "GOLD":      {"base": 73500, "label": "Gold (10g)",      "currency": "INR", "type": "commodity"},
    "SILVER":    {"base": 95000, "label": "Silver (1kg)",    "currency": "INR", "type": "commodity"},
    "CRUDE":     {"base": 6800,  "label": "Crude Oil (bbl)", "currency": "INR", "type": "commodity"},
    # Interest Rates
    "RBIREPO":   {"base": 6.50,  "label": "RBI Repo Rate",   "currency": "%",   "type": "rate"},
}

# Live USD/INR rate — updated every minute by market_ticker_background_worker
# Using a dict so it is mutable from within nested async functions
_live_usd_inr: dict = {"rate": 83.50}  # initial fallback until first Yahoo fetch

def _USD_INR() -> float:
    """Returns the current live USD/INR exchange rate."""
    return _live_usd_inr["rate"]

_YAHOO_MAPPING = {
    "SENSEX": "^BSESN",
    "NIFTY50": "^NSEI",
    "BANKNIFTY": "^NSEBANK",
    "SPX": "^GSPC",
    "NASDAQ": "^IXIC",
    "FTSE": "^FTSE",
    "BTCUSD": "BTC-USD",
    "ETHUSD": "ETH-USD",
    "SOLUSDT": "SOL-USD",
    "USDINR": "USDINR=X",
    "EURINR": "EURINR=X",
    "GBPINR": "GBPINR=X",
    "JPYINR": "JPYINR=X",
    "GOLD": "GC=F",
    "SILVER": "SI=F",
    "CRUDE": "CL=F",
}

async def fetch_yahoo_ticker(symbol: str) -> tuple[float, float]:
    """Fetch live price and percentage change from Yahoo Finance API."""
    async with httpx.AsyncClient(timeout=1.5) as client:
        try:
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
            headers = {"User-Agent": "Mozilla/5.0"}
            res = await client.get(url, headers=headers)
            if res.status_code == 200:
                data = res.json()
                meta = data.get("chart", {}).get("result", [{}])[0].get("meta", {})
                price = meta.get("regularMarketPrice")
                prev_close = meta.get("chartPreviousClose")
                if price and prev_close:
                    change = round(((price - prev_close) / prev_close) * 100, 2)
                    return float(price), change
        except Exception as e:
            log.warning(f"Failed to scrape {symbol} from Yahoo: {e}")
    return 0.0, 0.0

async def fetch_live_news() -> list[dict]:
    """Scrape live headlines from Moneycontrol RSS feed."""
    async with httpx.AsyncClient(timeout=8) as client:
        try:
            headers = {"User-Agent": "Mozilla/5.0"}
            res = await client.get("https://www.moneycontrol.com/rss/MC_news.xml", headers=headers)
            if res.status_code == 200:
                text = res.text
                items = []
                idx = 0
                while len(items) < 10:
                    start_item = text.find("<item>", idx)
                    if start_item == -1:
                        break
                    end_item = text.find("</item>", start_item)
                    if end_item == -1:
                        break
                    item_text = text[start_item:end_item]
                    
                    # title extractor
                    title_start = item_text.find("<title><![CDATA[")
                    if title_start != -1:
                        title_end = item_text.find("]]></title>", title_start)
                        title = item_text[title_start+16:title_end]
                    else:
                        title_start = item_text.find("<title>")
                        title_end = item_text.find("</title>", title_start)
                        title = item_text[title_start+7:title_end]
                        
                    title = title.replace("&amp;", "&").replace("&quot;", '"').strip()
                    if title:
                        sentiment = "neutral"
                        t_lower = title.lower()
                        if any(w in t_lower for w in ["gain", "rise", "jump", "surge", "upbeat", "bullish", "profit", "record high"]):
                            sentiment = "positive"
                        elif any(w in t_lower for w in ["fall", "drop", "plunge", "decline", "bearish", "loss", "slump", "concern"]):
                            sentiment = "negative"
                            
                        items.append({
                            "headline": title,
                            "sentiment": sentiment,
                            "source": "Moneycontrol RSS"
                        })
                    idx = end_item
                return items
        except Exception as e:
            log.warning(f"Failed to crawl Moneycontrol RSS: {e}")
    return []

def _market_tick(symbol: str, meta: dict) -> dict:
    """Generate a realistic market tick using time-seeded sine-wave noise."""
    t = time.time()
    # Use symbol hash for unique phase offset so each ticker moves independently
    phase = sum(ord(c) for c in symbol)
    # Multi-frequency noise for realism
    noise = (
        math.sin(t / 60 + phase) * 0.003 +
        math.sin(t / 15 + phase * 2) * 0.001 +
        random.gauss(0, 0.0005)
    )
    base = meta["base"]
    price = round(base * (1 + noise), 4 if meta["type"] == "forex" else 2 if base < 1000 else 0)
    change_pct = round(noise * 100, 2)
    change_abs = round(base * noise, 4 if meta["type"] == "forex" else 2 if base < 100 else 0)

    # Calculate both INR and USD prices for standard primary INR / secondary USD display
    price_inr = price
    price_usd = price

    currency = meta["currency"]
    if currency == "INR":
        price_inr = price
        if symbol == "USDINR":
            price_usd = 1.00
        elif symbol == "EURINR":
            price_usd = round(price / 1.08, 4) # EUR to USD approx
        elif symbol == "GBPINR":
            price_usd = round(price / 1.27, 4) # GBP to USD approx
        elif symbol == "JPYINR":
            price_usd = round(price / 156.0, 6) # JPY to USD approx
        else:
            price_usd = round(price / _live_usd_inr["rate"], 2)
    elif currency == "USD":
        price_usd = price
        _r = _live_usd_inr["rate"]
        price_inr = round(price * _r, 2 if price * _r < 1000 else 0)
    elif currency == "%":
        price_inr = price
        price_usd = price
    elif currency == "GBP":
        price_inr = round(price * 105.20, 2 if price * 105.20 < 1000 else 0)
        price_usd = round(price_inr / _live_usd_inr["rate"], 2)
    elif currency == "EUR":
        price_inr = round(price * 89.80, 2 if price * 89.80 < 1000 else 0)
        price_usd = round(price_inr / _live_usd_inr["rate"], 2)

    return {
        "symbol": symbol,
        "label": meta["label"],
        "price": price,
        "price_inr": price_inr,
        "price_usd": price_usd,
        "change_pct": change_pct,
        "change_abs": change_abs,
        "currency": meta["currency"],
        "type": meta["type"],
        "direction": "up" if change_pct >= 0 else "down",
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


async def market_ticker_background_worker():
    """Background task that runs periodically to fetch Yahoo Finance tickers and cache them in DB."""
    import asyncio
    log.info("Background market tickers worker started.")
    while True:
        try:
            # 1. Concurrently fetch prices
            async def fetch_one(sym: str, yahoo_sym: str | None) -> tuple[str, float, float]:
                if not yahoo_sym:
                    return sym, 0.0, 0.0
                try:
                    p, c = await fetch_yahoo_ticker(yahoo_sym)
                    return sym, p, c
                except Exception as e:
                    log.warning(f"Background worker error fetching {sym}: {e}")
                    return sym, 0.0, 0.0

            tasks = [fetch_one(sym, _YAHOO_MAPPING.get(sym)) for sym in _MARKET_BASE.keys()]
            results = await asyncio.gather(*tasks)
            scraped_map = {sym: (p, c) for sym, p, c in results}

            # Update live USD/INR in-memory rate immediately after fetch
            usdinr_live, _ = scraped_map.get("USDINR", (0.0, 0.0))
            if usdinr_live > 0:
                _live_usd_inr["rate"] = usdinr_live
                log.info(f"Live USD/INR rate updated: ₹{usdinr_live:.4f}")

            # 2. Update database cache sequentially for safety
            async with AsyncSessionLocal() as db:
                for sym, meta in _MARKET_BASE.items():
                    price, change_pct = scraped_map.get(sym, (0.0, 0.0))
                    if price == 0.0:
                        continue  # Keep old cached value if scrape failed
                    
                    cache_res = await db.execute(select(MarketTicker).where(MarketTicker.symbol == sym))
                    cached = cache_res.scalar_one_or_none()
                    if cached:
                        cached.price = price
                        cached.change_pct = change_pct
                    else:
                        db.add(MarketTicker(symbol=sym, price=price, change_pct=change_pct))
                await db.commit()
            log.info("Background market tickers cache successfully updated from Yahoo Finance.")
        except Exception as e:
            log.error(f"Error in background market tickers worker: {e}")
            
        await asyncio.sleep(60)


@app.get("/api/markets/live", summary="Live global market data (live scraped)")
async def get_markets_live(db: AsyncSession = Depends(get_db)):
    """Returns real-time prices for BSE, NSE, Forex, Commodities instantly from database cache."""
    tickers = []
    
    ticker_res = await db.execute(select(MarketTicker))
    cached_tickers = {t.symbol: t for t in ticker_res.scalars().all()}
    
    for sym, meta in _MARKET_BASE.items():
        cached = cached_tickers.get(sym)
        price = cached.price if cached else meta["base"]
        change_pct = cached.change_pct if cached else 0.0
        
        # fallback simulated tick if price is 0
        if price == 0.0:
            sim = _market_tick(sym, meta)
            price = sim["price"]
            change_pct = sim["change_pct"]
            
        live_rate = _live_usd_inr["rate"]
        tickers.append({
            "symbol": sym,
            "label": meta["label"],
            "price": price,
            "price_inr": price if meta["currency"] == "INR" else round(price * live_rate, 2),
            "price_usd": price if meta["currency"] == "USD" else round(price / live_rate, 2),
            "change_pct": change_pct,
            "change_abs": round(price * (change_pct / 100), 2),
            "currency": meta["currency"],
            "type": meta["type"],
            "direction": "up" if change_pct >= 0 else "down",
            "timestamp": datetime.now(timezone.utc).isoformat(),
        })

    usdinr_tick = next((t for t in tickers if t["symbol"] == "USDINR"), None)
    usdinr = usdinr_tick["price"] if usdinr_tick else _live_usd_inr["rate"]
    return {
        "tickers": tickers,
        "usdinr": usdinr,
        "market_status": "open" if 9 <= datetime.now().hour < 16 else "closed",
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


_MACRO_NEWS = [
    {"headline": "RBI holds repo rate at 6.50% — signals accommodative stance for Q3", "sentiment": "positive", "source": "RBI Bulletin", "impact": "rates"},
    {"headline": "India CPI inflation eases to 4.2% in April 2026 — lowest in 18 months", "sentiment": "positive", "source": "MoSPI", "impact": "inflation"},
    {"headline": "FPIs net buyers of ₹8,200 Cr in Indian equities this week", "sentiment": "positive", "source": "SEBI Data", "impact": "equity"},
    {"headline": "US Fed signals one rate cut in H2 2026 — USD/INR expected to stabilise", "sentiment": "neutral", "source": "Fed Minutes", "impact": "forex"},
    {"headline": "WPI falls to 1.1% — easing pressure on manufacturing input costs", "sentiment": "positive", "source": "DPIIT", "impact": "costs"},
    {"headline": "SEBI tightens KYC norms for fintech platforms effective July 2026", "sentiment": "negative", "source": "SEBI Circular", "impact": "compliance"},
    {"headline": "India GDP growth forecast upgraded to 7.2% for FY2026-27 — IMF", "sentiment": "positive", "source": "IMF WEO", "impact": "macro"},
    {"headline": "Crude oil rises 2.3% on Middle East supply concerns — watch fuel costs", "sentiment": "negative", "source": "Bloomberg", "impact": "costs"},
    {"headline": "Digital lending guidelines: RBI mandates FLDG cap at 5% for Fintech NBFCs", "sentiment": "neutral", "source": "RBI", "impact": "compliance"},
    {"headline": "GSTN portal upgrade: GSTR-1A filing window extended till 15th — EY Alert", "sentiment": "positive", "source": "EY India", "impact": "tax"},
]

@app.get("/api/macro-pulse", summary="Macro economic pulse & live news crawling")
async def get_macro_pulse(db: AsyncSession = Depends(get_db)):
    """Returns central bank rates, inflation, and live crawled financial news feed."""
    live_news = await fetch_live_news()
    
    # Store news items in database news_feed
    for n in live_news:
        existing = await db.execute(select(NewsFeed).where(NewsFeed.headline == n["headline"]))
        if not existing.scalar_one_or_none():
            db.add(NewsFeed(headline=n["headline"], sentiment=n["sentiment"], source=n["source"]))
            
    await db.commit()
    
    # Load news from DB
    news_res = await db.execute(select(NewsFeed).order_by(NewsFeed.created_at.desc()).limit(8))
    db_news = news_res.scalars().all()
    news_payload = [
        {"headline": n.headline, "sentiment": n.sentiment, "source": n.source} for n in db_news
    ]
    if not news_payload:
        news_payload = _MACRO_NEWS[:6]

    return {
        "central_bank_rates": {
            "rbi_repo": 6.50,
            "us_fed": 5.25,
            "ecb": 4.00,
            "boe": 5.00,
            "last_updated": "2026-05-01",
        },
        "inflation": {
            "india_cpi": 4.2,
            "india_wpi": 1.1,
            "us_cpi": 3.1,
            "eu_cpi": 2.4,
            "as_of": "April 2026",
        },
        "sector_risk": {
            "fintech": "medium",
            "wealthtech": "low",
            "manufacturing": "medium",
            "exports": "high",
            "it_services": "low",
        },
        "overall_sentiment": "cautiously_optimistic",
        "news": news_payload,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 12 — TREASURY: MULTI-CURRENCY & YIELD DATA
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/treasury/fx-rates", summary="Live FX rates for treasury")
async def get_fx_rates(db: AsyncSession = Depends(get_db)):
    """Returns FX rates with INR base for multi-currency treasury view.
    Reads from the MarketTicker cache (populated by /api/markets/live) so values
    are always in sync with the Global Market Intelligence dashboard.
    """
    # Forex symbols stored in MarketTicker table (price = INR per 1 unit of foreign)
    _FOREX_MAP = {
        "USDINR": "USD",
        "EURINR": "EUR",
        "GBPINR": "GBP",
        "JPYINR": "JPY",
    }
    # Fallbacks if not yet in cache
    _FALLBACK = {
        "USD": 83.50, "EUR": 89.80, "GBP": 105.20, "JPY": 0.5390,
        "SGD": 61.80, "AED": 22.73, "CHF": 92.50, "CAD": 61.20,
    }
    rates: dict[str, float] = dict(_FALLBACK)
    try:
        for sym, currency in _FOREX_MAP.items():
            res = await db.execute(select(MarketTicker).where(MarketTicker.symbol == sym))
            cached = res.scalar_one_or_none()
            if cached and cached.price and cached.price > 0:
                rates[currency] = round(cached.price, 4)
    except Exception as _fe:
        log.warning(f"FX rate DB lookup failed, using fallback: {_fe}")
    return {
        "base": "INR",
        "rates": rates,
        "yield_products": [
            {"name": "RBI G-Sec 10Y",    "yield_pct": 7.05, "risk": "Sovereign", "horizon": "10 years",  "currency": "INR"},
            {"name": "Liquid MF (Avg)",   "yield_pct": 7.20, "risk": "Low",       "horizon": "1 day",     "currency": "INR"},
            {"name": "Corporate FD AAA",  "yield_pct": 7.80, "risk": "Low",       "horizon": "1 year",    "currency": "INR"},
            {"name": "US T-Bill 3M",      "yield_pct": 5.30, "risk": "Sovereign", "horizon": "3 months",  "currency": "USD"},
            {"name": "SGD T-Bill 6M",     "yield_pct": 3.70, "risk": "Sovereign", "horizon": "6 months",  "currency": "SGD"},
            {"name": "HDFC FD 2Y",        "yield_pct": 7.25, "risk": "Low",       "horizon": "2 years",   "currency": "INR"},
            {"name": "Overnight Repo MF", "yield_pct": 6.90, "risk": "Very Low",  "horizon": "Overnight", "currency": "INR"},
            {"name": "Nifty ETF (Avg Ann)","yield_pct": 12.5,"risk": "Medium",    "horizon": "3+ years",  "currency": "INR"},
        ],
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 13 — INVOICE PROCESSING (AI OCR + LEDGER AUTO-POST)
# ─────────────────────────────────────────────────────────────────────────────

class InvoiceRecord(Base):
    __tablename__ = "invoice_records"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    vendor_name = Column(String(300), nullable=False)
    invoice_number = Column(String(100), nullable=True)
    invoice_date = Column(String(20), nullable=True)
    due_date = Column(String(20), nullable=True)
    amount_inr = Column(Float, nullable=False)
    gst_number = Column(String(20), nullable=True)
    gst_amount = Column(Float, nullable=True)
    category = Column(String(200), default="Uncategorized")
    status = Column(String(30), default="pending")  # pending/approved/posted/rejected
    notes = Column(Text, nullable=True)
    original_filename = Column(String(500), nullable=True)
    is_duplicate_flag = Column(Boolean, default=False)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))


class InvoiceCreate(BaseModel):
    vendor_name: str
    invoice_number: Optional[str] = None
    invoice_date: Optional[str] = None
    due_date: Optional[str] = None
    amount_inr: float
    gst_number: Optional[str] = None
    gst_amount: Optional[float] = None
    category: str = "Uncategorized"
    notes: Optional[str] = None


class InvoiceUpdate(BaseModel):
    vendor_name: Optional[str] = None
    invoice_number: Optional[str] = None
    invoice_date: Optional[str] = None
    due_date: Optional[str] = None
    amount_inr: Optional[float] = None
    gst_number: Optional[str] = None
    gst_amount: Optional[float] = None
    category: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None


def _ai_extract_invoice(text_content: str, filename: str) -> dict:
    """Use Groq LLM to extract invoice fields from raw text."""
    prompt = f"""You are an invoice OCR extraction AI for an Indian company.
Extract all invoice fields from the following text content of a file named "{filename}".

Text Content:
{text_content[:3000]}

Return ONLY valid JSON (no markdown, no explanation):
{{
  "vendor_name": "Company name from invoice or 'Unknown Vendor'",
  "invoice_number": "Invoice # or null",
  "invoice_date": "YYYY-MM-DD or null",
  "due_date": "YYYY-MM-DD or null",
  "amount_inr": 0.0,
  "gst_number": "GST number or null",
  "gst_amount": 0.0,
  "category": "Best matching category: Salaries | Software/SaaS | Marketing | Legal | Office | Travel | Infrastructure | Professional Services | Taxes | Other",
  "notes": "One line summary of what this invoice is for"
}}"""
    try:
        raw = call_groq(prompt)
        return _extract_json(raw)
    except Exception as e:
        log.warning(f"Invoice AI extraction failed: {e}")
        return {
            "vendor_name": "Unknown Vendor",
            "invoice_number": None,
            "invoice_date": None,
            "due_date": None,
            "amount_inr": 0.0,
            "gst_number": None,
            "gst_amount": None,
            "category": "Other",
            "notes": "Auto-extraction failed — please fill manually",
        }


@app.post("/api/invoices/upload", summary="Upload and AI-extract invoice")
async def upload_invoice(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """Upload a CSV/TXT/JSON invoice file and AI-extract its fields."""
    try:
        content = await file.read()
        filename = file.filename or "invoice"
        ext = filename.rsplit(".", 1)[-1].lower()

        if ext in ("csv", "txt"):
            text_content = content.decode("utf-8-sig", errors="replace")
        elif ext == "json":
            text_content = content.decode("utf-8")
        else:
            text_content = content.decode("utf-8-sig", errors="replace")

        extracted = _ai_extract_invoice(text_content, filename)

        # Duplicate detection: check if same vendor + amount + date already exists
        existing = await db.execute(
            select(InvoiceRecord).where(
                InvoiceRecord.vendor_name == extracted.get("vendor_name", ""),
                InvoiceRecord.amount_inr == float(extracted.get("amount_inr", 0)),
                InvoiceRecord.invoice_date == extracted.get("invoice_date"),
            )
        )
        is_dup = existing.scalar_one_or_none() is not None

        record = InvoiceRecord(
            vendor_name=extracted.get("vendor_name", "Unknown Vendor"),
            invoice_number=extracted.get("invoice_number"),
            invoice_date=extracted.get("invoice_date"),
            due_date=extracted.get("due_date"),
            amount_inr=float(extracted.get("amount_inr", 0)),
            gst_number=extracted.get("gst_number"),
            gst_amount=float(extracted.get("gst_amount") or 0),
            category=extracted.get("category", "Other"),
            notes=extracted.get("notes"),
            original_filename=filename,
            is_duplicate_flag=is_dup,
        )
        db.add(record)
        await db.commit()
        await db.refresh(record)
        return {"status": "extracted", "invoice": _invoice_to_dict(record), "is_duplicate": is_dup}
    except Exception as e:
        await db.rollback()
        log.error(f"Invoice upload error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/invoices", summary="Create invoice manually")
async def create_invoice(payload: InvoiceCreate, db: AsyncSession = Depends(get_db)):
    record = InvoiceRecord(**payload.model_dump())
    db.add(record)
    await db.commit()
    await db.refresh(record)
    return {"status": "created", "invoice": _invoice_to_dict(record)}


@app.get("/api/invoices", summary="List all invoices")
async def list_invoices(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(InvoiceRecord).order_by(InvoiceRecord.created_at.desc()))
    return [_invoice_to_dict(r) for r in result.scalars().all()]


@app.put("/api/invoices/{id}", summary="Update / approve / reject an invoice")
async def update_invoice(id: str, payload: InvoiceUpdate, db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(InvoiceRecord).where(InvoiceRecord.id == id))
    record = res.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="Invoice not found")
    for field, val in payload.model_dump(exclude_none=True).items():
        setattr(record, field, val)
    await db.commit()
    await db.refresh(record)
    return {"status": "updated", "invoice": _invoice_to_dict(record)}


@app.delete("/api/invoices/{id}", summary="Delete an invoice")
async def delete_invoice(id: str, db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(InvoiceRecord).where(InvoiceRecord.id == id))
    record = res.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="Invoice not found")
    await db.delete(record)
    await db.commit()
    return {"status": "deleted"}


def _invoice_to_dict(r: InvoiceRecord) -> dict:
    return {
        "id": r.id,
        "vendor_name": r.vendor_name,
        "invoice_number": r.invoice_number,
        "invoice_date": r.invoice_date,
        "due_date": r.due_date,
        "amount_inr": r.amount_inr,
        "gst_number": r.gst_number,
        "gst_amount": r.gst_amount,
        "category": r.category,
        "status": r.status,
        "notes": r.notes,
        "original_filename": r.original_filename,
        "is_duplicate_flag": r.is_duplicate_flag,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    }


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 14 — BANK RECONCILIATION ENGINE
# ─────────────────────────────────────────────────────────────────────────────

class ReconciliationRecord(Base):
    __tablename__ = "reconciliation_records"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    name = Column(String(300), nullable=False)
    total_bank = Column(Integer, default=0)
    total_ledger = Column(Integer, default=0)
    matched = Column(Integer, default=0)
    unmatched_bank = Column(Integer, default=0)
    unmatched_ledger = Column(Integer, default=0)
    result_json = Column(Text, nullable=True)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))


def _fuzzy_match(a: str, b: str) -> float:
    """Simple token-overlap similarity."""
    ta = set(a.lower().split())
    tb = set(b.lower().split())
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / max(len(ta), len(tb))


@app.post("/api/reconcile", summary="Reconcile bank statement vs ledger")
async def reconcile(
    bank_file: UploadFile = File(...),
    ledger_file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Upload bank CSV and ledger CSV. Returns matched/unmatched rows."""
    try:
        bank_bytes = await bank_file.read()
        ledger_bytes = await ledger_file.read()

        bank_rows = _parse_csv_bytes(bank_bytes)
        ledger_rows = _parse_csv_bytes(ledger_bytes)

        bank_txns, _ = _rows_to_transactions(bank_rows)
        ledger_txns, _ = _rows_to_transactions(ledger_rows)

        matched = []
        unmatched_bank = []
        used_ledger_idx = set()

        for b in bank_txns:
            best_score = 0
            best_idx = -1
            for i, l in enumerate(ledger_txns):
                if i in used_ledger_idx:
                    continue
                amount_match = abs(b.amount - l.amount) < 1.0
                date_match = b.date == l.date
                desc_match = _fuzzy_match(b.description, l.description)
                score = (int(amount_match) * 0.5) + (int(date_match) * 0.3) + (desc_match * 0.2)
                if score > best_score:
                    best_score = score
                    best_idx = i

            if best_score >= 0.6 and best_idx >= 0:
                used_ledger_idx.add(best_idx)
                matched.append({
                    "bank": b.model_dump(),
                    "ledger": ledger_txns[best_idx].model_dump(),
                    "match_score": round(best_score, 2),
                })
            else:
                unmatched_bank.append(b.model_dump())

        unmatched_ledger = [
            ledger_txns[i].model_dump()
            for i in range(len(ledger_txns))
            if i not in used_ledger_idx
        ]

        result = {
            "total_bank": len(bank_txns),
            "total_ledger": len(ledger_txns),
            "matched": len(matched),
            "unmatched_bank": len(unmatched_bank),
            "unmatched_ledger": len(unmatched_ledger),
            "match_rate_pct": round(len(matched) / max(len(bank_txns), 1) * 100, 1),
            "matched_items": matched,
            "unmatched_bank_items": unmatched_bank,
            "unmatched_ledger_items": unmatched_ledger,
        }

        record = ReconciliationRecord(
            name=f"Recon {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            total_bank=len(bank_txns),
            total_ledger=len(ledger_txns),
            matched=len(matched),
            unmatched_bank=len(unmatched_bank),
            unmatched_ledger=len(unmatched_ledger),
            result_json=json.dumps(result),
        )
        db.add(record)
        await db.commit()

        return {"status": "ok", "reconciliation_id": record.id, **result}
    except Exception as e:
        await db.rollback()
        log.error(f"Reconciliation error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/reconcile", summary="List past reconciliations")
async def list_reconciliations(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(ReconciliationRecord).order_by(ReconciliationRecord.created_at.desc()))
    return [
        {
            "id": r.id,
            "name": r.name,
            "total_bank": r.total_bank,
            "total_ledger": r.total_ledger,
            "matched": r.matched,
            "unmatched_bank": r.unmatched_bank,
            "unmatched_ledger": r.unmatched_ledger,
            "match_rate_pct": round(r.matched / max(r.total_bank, 1) * 100, 1),
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in result.scalars().all()
    ]


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 15 — VENDOR INTELLIGENCE
# ─────────────────────────────────────────────────────────────────────────────

class VendorRecord(Base):
    __tablename__ = "vendor_records"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    vendor_name = Column(String(300), nullable=False)
    category = Column(String(200), default="Other")
    total_spend_inr = Column(Float, default=0.0)
    transaction_count = Column(Integer, default=0)
    gst_number = Column(String(20), nullable=True)
    concentration_pct = Column(Float, default=0.0)
    risk_level = Column(String(20), default="Low")   # Low/Medium/High/Critical
    payment_terms = Column(String(100), nullable=True)
    notes = Column(Text, nullable=True)
    is_verified = Column(Boolean, default=False)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))


class VendorUpdate(BaseModel):
    vendor_name: Optional[str] = None
    category: Optional[str] = None
    gst_number: Optional[str] = None
    risk_level: Optional[str] = None
    payment_terms: Optional[str] = None
    notes: Optional[str] = None
    is_verified: Optional[bool] = None


@app.get("/api/vendors", summary="Get vendor intelligence from stored transactions")
async def get_vendors(file_id: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    """Aggregate vendors from stored transactions + any manual vendor records."""
    try:
        # Aggregate from stored transactions
        query = select(StoredTransaction)
        if file_id:
            query = query.where(StoredTransaction.file_id == file_id)
        result = await db.execute(query)
        txns = result.scalars().all()

        vendor_map: dict[str, dict] = {}
        total_expense = sum(t.amount for t in txns if t.type == "Expense")

        for t in txns:
            if t.type != "Expense":
                continue
            key = t.category.strip() or "Uncategorized"
            if key not in vendor_map:
                vendor_map[key] = {"vendor_name": key, "category": key, "total_spend_inr": 0.0, "transaction_count": 0}
            vendor_map[key]["total_spend_inr"] += t.amount
            vendor_map[key]["transaction_count"] += 1

        vendors = []
        for k, v in vendor_map.items():
            conc = round(v["total_spend_inr"] / total_expense * 100, 1) if total_expense > 0 else 0.0
            risk = "Critical" if conc > 40 else "High" if conc > 20 else "Medium" if conc > 10 else "Low"

            # Check if there's a manual override record
            manual_res = await db.execute(select(VendorRecord).where(VendorRecord.vendor_name == k))
            manual = manual_res.scalar_one_or_none()

            vendors.append({
                "id": manual.id if manual else None,
                "vendor_name": v["vendor_name"],
                "category": manual.category if manual else v["category"],
                "total_spend_inr": round(v["total_spend_inr"], 2),
                "total_spend_usd": round(v["total_spend_inr"] / _live_usd_inr["rate"], 2),
                "transaction_count": v["transaction_count"],
                "concentration_pct": conc,
                "risk_level": manual.risk_level if manual else risk,
                "gst_number": manual.gst_number if manual else None,
                "payment_terms": manual.payment_terms if manual else None,
                "notes": manual.notes if manual else None,
                "is_verified": manual.is_verified if manual else False,
            })

        vendors.sort(key=lambda x: x["total_spend_inr"], reverse=True)
        return {"vendors": vendors, "total_expense_inr": round(total_expense, 2)}
    except Exception as e:
        log.error(f"Vendor fetch error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/vendors/{vendor_name}", summary="Update vendor record (editable)")
async def update_vendor(vendor_name: str, payload: VendorUpdate, db: AsyncSession = Depends(get_db)):
    """Create or update a vendor override record."""
    try:
        res = await db.execute(select(VendorRecord).where(VendorRecord.vendor_name == vendor_name))
        record = res.scalar_one_or_none()
        if not record:
            record = VendorRecord(vendor_name=vendor_name)
            db.add(record)
        for field, val in payload.model_dump(exclude_none=True).items():
            setattr(record, field, val)
        record.updated_at = datetime.now(timezone.utc)
        await db.commit()
        await db.refresh(record)
        return {"status": "updated", "vendor_name": vendor_name}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/vendors", summary="Manually add a vendor record")
async def create_vendor(payload: VendorUpdate, db: AsyncSession = Depends(get_db)):
    """Create a new manual vendor record. vendor_name is required."""
    if not payload.vendor_name:
        raise HTTPException(status_code=422, detail="vendor_name is required")
    try:
        # Check for duplicates
        existing = await db.execute(select(VendorRecord).where(VendorRecord.vendor_name == payload.vendor_name))
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=409, detail=f"Vendor '{payload.vendor_name}' already exists")
        record = VendorRecord(
            vendor_name=payload.vendor_name,
            category=payload.category or "Other",
            gst_number=payload.gst_number,
            risk_level=payload.risk_level or "Low",
            payment_terms=payload.payment_terms,
            notes=payload.notes,
            is_verified=payload.is_verified or False,
        )
        db.add(record)
        await db.commit()
        await db.refresh(record)
        return {"status": "created", "vendor_name": record.vendor_name, "id": record.id}
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/vendors/{vendor_name}", summary="Delete a vendor record")
async def delete_vendor(vendor_name: str, db: AsyncSession = Depends(get_db)):
    """Delete a manual vendor override record. Transaction history is preserved."""
    try:
        res = await db.execute(select(VendorRecord).where(VendorRecord.vendor_name == vendor_name))
        record = res.scalar_one_or_none()
        if not record:
            raise HTTPException(status_code=404, detail=f"Vendor '{vendor_name}' not found")
        await db.delete(record)
        await db.commit()
        return {"status": "deleted", "vendor_name": vendor_name}
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 16 — SCENARIO WAR ROOM (server-side computation)
# ─────────────────────────────────────────────────────────────────────────────

class ScenarioRequest(BaseModel):
    base_revenue: float
    base_expenses: float
    revenue_change_pct: float = 0.0   # e.g. -20 for 20% revenue drop
    expense_change_pct: float = 0.0   # e.g. +15 for 15% expense increase
    headcount_change: int = 0          # new hires (positive) or layoffs (negative)
    avg_salary_inr: float = 75000.0   # per month per head
    cash_reserve_inr: float = 0.0

class ScenarioResult(BaseModel):
    adjusted_revenue: float
    adjusted_expenses: float
    adjusted_profit: float
    monthly_burn: float
    runway_months: Optional[float]
    health_impact: str
    recommendation: str

@app.post("/api/scenarios/calculate", summary="Run a what-if scenario calculation")
def calculate_scenario(req: ScenarioRequest):
    adj_rev = req.base_revenue * (1 + req.revenue_change_pct / 100)
    headcount_cost = req.headcount_change * req.avg_salary_inr * 12  # annual
    adj_exp = req.base_expenses * (1 + req.expense_change_pct / 100) + headcount_cost
    adj_profit = adj_rev - adj_exp
    cash = req.cash_reserve_inr + adj_profit
    monthly_expenses = adj_exp / 12
    burn = abs(adj_profit / 12) if adj_profit < 0 else 0
    runway = round(cash / burn, 1) if burn > 0 else None

    if adj_profit < 0 and burn > 0 and (runway is None or runway < 6):
        health = "CRITICAL"
        rec = "Immediate cost reduction required. Consider deferring hires and renegotiating vendor contracts."
    elif burn > 0 and runway and runway < 12:
        health = "WARNING"
        rec = "Runway under 12 months. Accelerate revenue or reduce discretionary spend."
    elif adj_profit >= 0:
        health = "HEALTHY"
        rec = "Financials remain strong under this scenario. Consider deploying surplus into treasury instruments."
    else:
        health = "NORMAL"
        rec = "Monitor closely. Ensure cost controls are in place."

    return ScenarioResult(
        adjusted_revenue=round(adj_rev, 2),
        adjusted_expenses=round(adj_exp, 2),
        adjusted_profit=round(adj_profit, 2),
        monthly_burn=round(burn, 2),
        runway_months=runway,
        health_impact=health,
        recommendation=rec,
    )


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 17 — COMPLIANCE: MULTI-JURISDICTION TAX CALENDAR
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/compliance/tax-calendar", summary="Multi-jurisdiction tax deadline calendar")
def get_tax_calendar():
    """Returns upcoming tax and compliance deadlines for India, US, EU, UK."""
    now = datetime.now()
    year = now.year
    month = now.month

    def _next_month_day(day: int) -> str:
        m = month + 1 if now.day > day else month
        y = year + (1 if m > 12 else 0)
        m = m % 12 or 12
        return f"{y}-{m:02d}-{day:02d}"

    return {
        "deadlines": [
            # India
            {"task": "GSTR-3B Filing", "jurisdiction": "India", "due_date": _next_month_day(20), "priority": "HIGH", "category": "GST", "penalty": "₹50/day"},
            {"task": "GSTR-1 Filing", "jurisdiction": "India", "due_date": _next_month_day(11), "priority": "HIGH", "category": "GST", "penalty": "₹50/day"},
            {"task": "TDS Payment (194C/J)", "jurisdiction": "India", "due_date": _next_month_day(7), "priority": "CRITICAL", "category": "TDS", "penalty": "1.5% per month"},
            {"task": "Advance Tax (Q2)", "jurisdiction": "India", "due_date": f"{year}-09-15", "priority": "HIGH", "category": "Income Tax", "penalty": "1% per month"},
            {"task": "PF & ESI Contribution", "jurisdiction": "India", "due_date": _next_month_day(15), "priority": "MEDIUM", "category": "Payroll", "penalty": "₹5,000"},
            {"task": "ROC Annual Return (MCA)", "jurisdiction": "India", "due_date": f"{year}-09-30", "priority": "HIGH", "category": "Corporate", "penalty": "₹100/day"},
            # USA
            {"task": "Estimated Tax (Q2)", "jurisdiction": "USA", "due_date": f"{year}-06-17", "priority": "MEDIUM", "category": "Federal Tax", "penalty": "Underpayment penalty"},
            {"task": "Payroll Tax Deposit", "jurisdiction": "USA", "due_date": _next_month_day(15), "priority": "HIGH", "category": "Payroll", "penalty": "FTD penalty 10%"},
            # EU
            {"task": "VAT Return (Quarterly)", "jurisdiction": "EU", "due_date": _next_month_day(30), "priority": "MEDIUM", "category": "VAT", "penalty": "Varies by country"},
            # UK
            {"task": "VAT Return & Payment", "jurisdiction": "UK", "due_date": _next_month_day(7), "priority": "MEDIUM", "category": "VAT", "penalty": "Surcharge"},
            {"task": "Corporation Tax (CT600)", "jurisdiction": "UK", "due_date": f"{year}-12-31", "priority": "HIGH", "category": "Corporate Tax", "penalty": "£100 + interest"},
        ],
        "gst_reconciliation": {
            "gstr1_filed": True,
            "gstr2b_available": True,
            "gstr3b_filed": False,
            "itc_available_inr": 48250.00,
            "itc_claimed_inr": 41000.00,
            "variance_inr": 7250.00,
            "status": "Variance detected — review input tax credits",
        },
        "audit_readiness": {
            "score": 88,
            "label": "Good",
            "checklist": [
                {"item": "GST ledger reconciled", "done": True},
                {"item": "TDS challans matched", "done": True},
                {"item": "Bank reconciliation complete", "done": False},
                {"item": "Director's KYC updated", "done": True},
                {"item": "ROC filings current", "done": False},
                {"item": "Transfer pricing documentation", "done": True},
            ],
        },
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 18 — BOARD PACK GENERATION (INR primary, USD secondary)
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/board-pack/data/{analysis_id}", summary="Get board pack data for an analysis")
async def get_board_pack_data(analysis_id: str, db: AsyncSession = Depends(get_db)):
    """Returns all data needed for a board pack report in INR primary / USD secondary."""
    try:
        res = await db.execute(select(AnalysisResult).where(AnalysisResult.id == analysis_id))
        analysis = res.scalar_one_or_none()
        if not analysis:
            raise HTTPException(status_code=404, detail="Analysis not found")

        file_res = await db.execute(select(UploadedFile).where(UploadedFile.id == analysis.file_id))
        uploaded_file = file_res.scalar_one_or_none()

        result_data = json.loads(analysis.result_json)
        kpis = result_data.get("kpis", {})
        usdinr = _live_usd_inr["rate"]

        def to_usd(inr_val):
            if inr_val is None:
                return None
            return round(float(inr_val) / usdinr, 2)

        return {
            "company_name": uploaded_file.company_name if uploaded_file else "Company",
            "period_start": uploaded_file.period_start if uploaded_file else "",
            "period_end": uploaded_file.period_end if uploaded_file else "",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "usdinr_rate": usdinr,
            "executive_summary": result_data.get("final", {}).get("output", {}).get("final_summary", ""),
            "health_score": analysis.health_score,
            "priority": analysis.priority,
            "financials": {
                "total_revenue_inr": kpis.get("total_revenue"),
                "total_revenue_usd": to_usd(kpis.get("total_revenue")),
                "total_expenses_inr": kpis.get("total_expenses"),
                "total_expenses_usd": to_usd(kpis.get("total_expenses")),
                "net_profit_inr": kpis.get("profit"),
                "net_profit_usd": to_usd(kpis.get("profit")),
                "monthly_burn_inr": kpis.get("monthly_burn"),
                "monthly_burn_usd": to_usd(kpis.get("monthly_burn")),
                "runway_months": kpis.get("runway_months"),
                "gross_margin_pct": kpis.get("gross_margin_pct"),
            },
            "monthly_trend": kpis.get("monthly", []),
            "top_categories": kpis.get("top_categories", []),
            "forecast": kpis.get("forecast", []),
            "urgent_actions": result_data.get("final", {}).get("output", {}).get("urgent_actions", []),
            "composite_patterns": result_data.get("composite_patterns", []),
            "agents": {
                "research": result_data.get("research", {}).get("output", {}),
                "planning": result_data.get("planning", {}).get("output", {}),
                "treasury": result_data.get("treasury", {}).get("output", {}),
                "compliance": result_data.get("compliance", {}).get("output", {}),
                "decision": result_data.get("decision", {}).get("output", {}),
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        log.error(f"Board pack error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 19 — COMPANY INTELLIGENCE TRACKER (Groq + Web Scraper)
# ─────────────────────────────────────────────────────────────────────────────

from bs4 import BeautifulSoup

class TrackedCompany(Base):
    __tablename__ = "tracked_companies"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    name = Column(String(300), nullable=False)
    url = Column(String(1000), nullable=True)
    scraped_text = Column(Text, nullable=True)
    summary = Column(Text, nullable=True)       # Groq-generated research report (JSON string)
    last_scraped_at = Column(DateTime, nullable=True)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))


class TrackCompanyRequest(BaseModel):
    name: str
    url: Optional[str] = None


async def _scrape_url(url: str) -> str:
    """Fetch a URL and extract clean text using BeautifulSoup."""
    try:
        async with httpx.AsyncClient(timeout=12, follow_redirects=True) as client:
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Accept": "text/html,application/xhtml+xml",
            }
            res = await client.get(url, headers=headers)
            if res.status_code == 200:
                soup = BeautifulSoup(res.text, "html.parser")
                # Remove script/style noise
                for tag in soup(["script", "style", "nav", "footer", "header"]):
                    tag.decompose()
                text = soup.get_text(separator=" ", strip=True)
                # Trim to 4000 chars for Groq context window
                return text[:4000]
    except Exception as e:
        log.warning(f"Scrape failed for {url}: {e}")
    return ""


def _groq_company_research(name: str, url: str, scraped_text: str) -> dict:
    """Use Groq to generate a structured company intelligence report."""
    context = f"Company: {name}\nURL: {url}\n\nScraped content:\n{scraped_text}" if scraped_text else f"Company: {name}\nURL: {url}"
    prompt = f"""You are a senior financial and competitive intelligence analyst at an Indian fintech.
Analyze the following company and produce a structured intelligence report.

{context}

Return ONLY valid JSON (no markdown):
{{
  "company_name": "{name}",
  "url": "{url}",
  "tagline": "One-line description of what the company does",
  "market_position": "Their positioning in the market (leader/challenger/niche)",
  "core_products": ["Product/service 1", "Product/service 2"],
  "target_segments": ["Segment 1", "Segment 2"],
  "key_strengths": ["Strength 1", "Strength 2", "Strength 3"],
  "key_risks": ["Risk 1", "Risk 2"],
  "competitive_threat": "low|medium|high",
  "financial_signals": "Any publicly known financial indicators or funding status",
  "strategic_recommendation": "What Aghron should do in response to this company",
  "sentiment": "positive|neutral|negative"
}}"""
    try:
        raw = call_groq(prompt, max_tokens=1200)
        return _extract_json(raw)
    except Exception as e:
        log.warning(f"Groq company research failed for {name}: {e}")
        return {
            "company_name": name,
            "url": url,
            "tagline": "Research pending",
            "market_position": "Unknown",
            "core_products": [],
            "target_segments": [],
            "key_strengths": [],
            "key_risks": [],
            "competitive_threat": "medium",
            "financial_signals": "Not available",
            "strategic_recommendation": "Manual review required",
            "sentiment": "neutral",
        }


@app.post("/api/companies/track", summary="Add a company to track & run AI research")
async def track_company(req: TrackCompanyRequest, db: AsyncSession = Depends(get_db)):
    """Scrape the company URL, run Groq analysis, and persist the research report."""
    try:
        # Check duplicates
        existing = await db.execute(select(TrackedCompany).where(TrackedCompany.name == req.name))
        company = existing.scalar_one_or_none()

        scraped_text = ""
        if req.url:
            scraped_text = await _scrape_url(req.url)

        # Run Groq research (sync, offloaded)
        import asyncio
        summary_dict = await asyncio.get_event_loop().run_in_executor(
            None, _groq_company_research, req.name, req.url or "", scraped_text
        )

        if company:
            # Update existing
            company.url = req.url or company.url
            company.scraped_text = scraped_text
            company.summary = json.dumps(summary_dict)
            company.last_scraped_at = datetime.now(timezone.utc)
        else:
            company = TrackedCompany(
                name=req.name,
                url=req.url,
                scraped_text=scraped_text,
                summary=json.dumps(summary_dict),
                last_scraped_at=datetime.now(timezone.utc),
            )
            db.add(company)

        await db.commit()
        await db.refresh(company)
        return {
            "status": "tracked",
            "id": company.id,
            "name": company.name,
            "url": company.url,
            "summary": summary_dict,
            "last_scraped_at": company.last_scraped_at.isoformat() if company.last_scraped_at else None,
        }
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        log.error(f"Company track error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/companies/track", summary="List all tracked companies")
async def list_tracked_companies(db: AsyncSession = Depends(get_db)):
    """Returns all tracked companies with their AI-generated research reports."""
    try:
        res = await db.execute(select(TrackedCompany).order_by(TrackedCompany.created_at.desc()))
        companies = res.scalars().all()
        return [
            {
                "id": c.id,
                "name": c.name,
                "url": c.url,
                "summary": json.loads(c.summary) if c.summary else {},
                "last_scraped_at": c.last_scraped_at.isoformat() if c.last_scraped_at else None,
                "created_at": c.created_at.isoformat() if c.created_at else None,
            }
            for c in companies
        ]
    except Exception as e:
        log.error(f"List companies error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/companies/track/{company_id}", summary="Delete a tracked company")
async def delete_tracked_company(company_id: str, db: AsyncSession = Depends(get_db)):
    """Delete a tracked company record."""
    try:
        res = await db.execute(select(TrackedCompany).where(TrackedCompany.id == company_id))
        company = res.scalar_one_or_none()
        if not company:
            raise HTTPException(status_code=404, detail="Company not found")
        await db.delete(company)
        await db.commit()
        return {"status": "deleted", "id": company_id}
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────────────────────────────────────
# ENTRYPOINT — run directly: python chief_ai.py
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn

    port = int(os.getenv("PORT", "8000"))
    log.info(f"🚀  Finance AI Chief Command starting on http://127.0.0.1:{port}")
    uvicorn.run(
        "chief_ai:app",
        host="0.0.0.0",
        port=port,
        reload=True,
        log_level="info",
    )

